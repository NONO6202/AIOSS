"TXT_REDACTED"
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "TXT_REDACTED" / "TXT_REDACTED"
DEFAULT_OUT = ROOT / "TXT_REDACTED" / "TXT_REDACTED"


def snapshot(xlsx_path: Path, sample_rows: int) -> dict:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    result: dict = {"TXT_REDACTED": str(xlsx_path), "TXT_REDACTED": []}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers: list = []
        samples: list[list] = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx == 2:
                headers = [c for c in row]
            else:
                samples.append([c for c in row])
            if idx >= sample_rows:
                break
        result["TXT_REDACTED"].append(
            {
                "TXT_REDACTED": sheet_name,
                "TXT_REDACTED": len(headers),
                "TXT_REDACTED": headers,
                "TXT_REDACTED": samples,
            }
        )
    wb.close()
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("TXT_REDACTED", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("TXT_REDACTED", type=int, default=3, help="TXT_REDACTED")
    parser.add_argument("TXT_REDACTED", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    snap = snapshot(args.xlsx, args.rows)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(
        json.dumps(snap, ensure_ascii=False, indent=4, default=str),
        encoding="TXT_REDACTED",
    )
    for sheet in snap["TXT_REDACTED"]:
        print("TXT_REDACTED"                                            )
    print("TXT_REDACTED"               )


if __name__ == "TXT_REDACTED":
    main()
