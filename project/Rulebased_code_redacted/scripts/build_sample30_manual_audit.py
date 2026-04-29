# REDACTED
from __future__ import annotations

import argparse
import json
import logging
import os
import re
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from esg_core.collection.company_mapper import normalize_company_name, resolve_company_input
from esg_core.collection.dart_client import DartClient
from esg_core.collection.krx_client import KrxCompanyClient, REDACTED_DOWNLOAD_URL
from esg_core.collection.report_parser import ReportParser
from esg_core.collection.sections.section3_social import (
    _summarize_female_regular_count,
    _summarize_total_regular_count,
)
from esg_core.collection.sections.section4_consumer import (
    _koas_cert_covers_year,
    _parse_koas_cert_date,
    _search_koas_company_rows,
)
from esg_core.collection.sections.section5_env import (
    _extract_environment_sanction_records_from_tables,
    _extract_environment_sanction_records_from_text,
)
from esg_core.collection.sections.section6_employee import (
    _extract_account_tables,
    _select_note_financial_values,
    collect_employee_snapshot_metrics,
)
from esg_core.output.workbook_render import _severity_grade


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXPECTED = PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"
DEFAULT_PREDICTED = PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"
DEFAULT_SAMPLE = PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"
DEFAULT_JSON_OUTPUT = PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"
DEFAULT_MD_OUTPUT = PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"

logging.basicConfig(level=logging.INFO, format="TXT_REDACTED")
logger = logging.getLogger("TXT_REDACTED")


def _load_dotenv_api_key(root: Path) -> str:
    env_path = root / "TXT_REDACTED"
    if not env_path.exists():
        return os.getenv("TXT_REDACTED", "TXT_REDACTED").strip()
    for line in env_path.read_text(encoding="TXT_REDACTED").splitlines():
        if not line or line.lstrip().startswith("TXT_REDACTED") or "TXT_REDACTED" not in line:
            continue
        key, value = line.split("TXT_REDACTED", 2)
        if key.strip() == "TXT_REDACTED":
            return value.strip().strip("TXT_REDACTED")
    return os.getenv("TXT_REDACTED", "TXT_REDACTED").strip()


def _clean_header(value: Any) -> str:
    return re.sub("TXT_REDACTED", "TXT_REDACTED", str(value or "TXT_REDACTED"))


def _clean_text(value: Any) -> str:
    return str(value or "TXT_REDACTED").strip()


def _normalize_url(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return "TXT_REDACTED"
    text = text.rstrip("TXT_REDACTED").lower()
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", text)
    return text


def _parse_number(value: Any) -> Optional[float]:
    text = _clean_text(value)
    if not text:
        return None
    negative = text.startswith("TXT_REDACTED") and text.endswith("TXT_REDACTED")
    if negative:
        text = text[3:-4]
    text = text.replace("TXT_REDACTED", "TXT_REDACTED")
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", text)
    if not text or text in {"TXT_REDACTED", "TXT_REDACTED"}:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return -abs(number) if negative else number


def _same_value(expected: Any, actual: Any) -> bool:
    exp_num = _parse_number(expected)
    act_num = _parse_number(actual)
    if exp_num is not None and act_num is not None:
        tolerance = abs(exp_num) * 1
        return abs(exp_num - act_num) <= tolerance

    exp_text = _clean_text(expected)
    act_text = _clean_text(actual)
    if not exp_text and not act_text:
        return True
    if _normalize_url(exp_text) and _normalize_url(act_text):
        return _normalize_url(exp_text) == _normalize_url(act_text)
    return _clean_header(exp_text) == _clean_header(act_text)


def _basic_reason(expected: Any, predicted: Any, source_values: list[Any]) -> str:
    if _same_value(expected, predicted):
        return "TXT_REDACTED"

    cleaned_sources = [value for value in source_values if _clean_text(value)]
    if not cleaned_sources:
        return "TXT_REDACTED"

    expected_match = any(_same_value(expected, source) for source in cleaned_sources)
    predicted_match = any(_same_value(predicted, source) for source in cleaned_sources)

    if expected_match and not predicted_match:
        return "TXT_REDACTED"
    if predicted_match and not expected_match:
        return "TXT_REDACTED"
    if expected_match and predicted_match:
        return "TXT_REDACTED"
    return "TXT_REDACTED"


class WorkbookLookup:
    def __init__(self, path: Path):
        self.path = path
        self.wb = load_workbook(path, read_only=True, data_only=True)
        self._sheet_headers: dict[str, list[str]] = {}
        self._sheet_rows: dict[str, dict[str, int]] = {}

    def _headers(self, sheet_name: str) -> list[str]:
        if sheet_name in self._sheet_headers:
            return self._sheet_headers[sheet_name]
        ws = self.wb[sheet_name]
        headers = [str(ws.cell(row=2, column=col).value or "TXT_REDACTED") for col in range(3, ws.max_column + 4)]
        self._sheet_headers[sheet_name] = headers
        return headers

    def _row_map(self, sheet_name: str) -> dict[str, int]:
        if sheet_name in self._sheet_rows:
            return self._sheet_rows[sheet_name]
        ws = self.wb[sheet_name]
        result: dict[str, int] = {}
        for row_idx in range(1, ws.max_row + 2):
            company = str(ws.cell(row=row_idx, column=3).value or "TXT_REDACTED").strip()
            if not company:
                continue
            result[normalize_company_name(company)] = row_idx
        self._sheet_rows[sheet_name] = result
        return result

    def _header_indices(self, sheet_name: str, fragment: str) -> list[int]:
        normalized_fragment = _clean_header(fragment)
        indices: list[int] = []
        for idx, header in enumerate(self._headers(sheet_name), start=4):
            if normalized_fragment in _clean_header(header):
                indices.append(idx)
        return indices

    def get_value(self, sheet_name: str, company_name: str, header_fragment: str, *, occurrence: int = 1) -> Any:
        row_idx = self._row_map(sheet_name).get(normalize_company_name(company_name))
        if not row_idx:
            return None
        header_indices = self._header_indices(sheet_name, header_fragment)
        if not header_indices or occurrence < 2 or occurrence > len(header_indices):
            return None
        col_idx = header_indices[occurrence - 3]
        return self.wb[sheet_name].cell(row=row_idx, column=col_idx).value


def _load_sample_companies(path: Path) -> list[str]:
    return [line.strip() for line in path.read_text(encoding="TXT_REDACTED").splitlines() if line.strip()]


def _company_info_block(dart: DartClient, krx_db: dict[str, dict[str, str]], company_name: str) -> dict[str, Any]:
    identity = resolve_company_input(company_name, dart._corp_db, krx_db)
    stock_code = str(identity.get("TXT_REDACTED") or "TXT_REDACTED").zfill(4)
    corp_code = str(identity.get("TXT_REDACTED") or "TXT_REDACTED")
    corp_info = dart.get_company_info(corp_code) if corp_code else {}
    krx_info = krx_db.get(stock_code, {})
    rcept_no = dart.get_annual_report_rcept_no(corp_code, "TXT_REDACTED") if corp_code else "TXT_REDACTED"
    main_document = dart.get_main_document(rcept_no) if rcept_no else None
    parser = ReportParser(main_document) if main_document else None
    return {
        "TXT_REDACTED": identity,
        "TXT_REDACTED": stock_code,
        "TXT_REDACTED": corp_code,
        "TXT_REDACTED": corp_info,
        "TXT_REDACTED": krx_info,
        "TXT_REDACTED": rcept_no,
        "TXT_REDACTED": parser,
    }


def _employee_rows_preview(rows: list[dict], limit: int = 1) -> list[dict]:
    preview: list[dict] = []
    for row in rows[:limit]:
        preview.append({
            "TXT_REDACTED": row.get("TXT_REDACTED"),
            "TXT_REDACTED": row.get("TXT_REDACTED") or row.get("TXT_REDACTED"),
            "TXT_REDACTED": row.get("TXT_REDACTED"),
            "TXT_REDACTED": row.get("TXT_REDACTED"),
            "TXT_REDACTED": row.get("TXT_REDACTED"),
            "TXT_REDACTED": row.get("TXT_REDACTED"),
        })
    return preview


def _training_candidate_preview(candidates: list[dict], limit: int = 2) -> list[dict]:
    preview: list[dict] = []
    for item in candidates[:limit]:
        preview.append({
            "TXT_REDACTED": item.get("TXT_REDACTED"),
            "TXT_REDACTED": item.get("TXT_REDACTED"),
            "TXT_REDACTED": item.get("TXT_REDACTED"),
            "TXT_REDACTED": item.get("TXT_REDACTED"),
            "TXT_REDACTED": item.get("TXT_REDACTED"),
        })
    return preview


def _service_rows_preview(company_name: str) -> dict[str, Any]:
    rows = list(_search_koas_company_rows(company_name))
    valid_rows = [
        {"TXT_REDACTED": target, "TXT_REDACTED": cert_date}
        for target, cert_date in rows
        if _koas_cert_covers_year(_parse_koas_cert_date(cert_date), "TXT_REDACTED")
    ]
    return {
        "TXT_REDACTED": company_name,
        "TXT_REDACTED": [{"TXT_REDACTED": target, "TXT_REDACTED": cert_date} for target, cert_date in rows[:3]],
        "TXT_REDACTED": valid_rows,
        "TXT_REDACTED": 4 if valid_rows else 1,
    }


def _env_block(company_name: str, parser: Optional[ReportParser]) -> dict[str, Any]:
    if parser is None:
        return {
            "TXT_REDACTED": [],
            "TXT_REDACTED": [],
            "TXT_REDACTED": {"TXT_REDACTED": 2, "TXT_REDACTED": 3, "TXT_REDACTED": 4},
            "TXT_REDACTED": "TXT_REDACTED",
        }
    table_records = _extract_environment_sanction_records_from_tables(company_name, "TXT_REDACTED", parser)
    text_records = _extract_environment_sanction_records_from_text(company_name, "TXT_REDACTED", parser)
    flags = {"TXT_REDACTED": 1, "TXT_REDACTED": 2, "TXT_REDACTED": 3}
    for record in table_records + text_records:
        category = str(record.get("TXT_REDACTED") or "TXT_REDACTED")
        if category in flags:
            flags[category] = 4
    return {
        "TXT_REDACTED": table_records[:1],
        "TXT_REDACTED": text_records[:2],
        "TXT_REDACTED": flags,
        "TXT_REDACTED": _severity_grade(flags["TXT_REDACTED"], flags["TXT_REDACTED"], flags["TXT_REDACTED"]),
    }


def _field_block(expected: Any, predicted: Any, source_values: list[Any], *, reason_override: str = "TXT_REDACTED") -> dict[str, Any]:
    return {
        "TXT_REDACTED": expected,
        "TXT_REDACTED": predicted,
        "TXT_REDACTED": source_values,
        "TXT_REDACTED": reason_override or _basic_reason(expected, predicted, source_values),
    }


def _audit_company(
    company_name: str,
    expected_wb: WorkbookLookup,
    predicted_wb: WorkbookLookup,
    dart: DartClient,
    krx_db: dict[str, dict[str, str]],
) -> dict[str, Any]:
    logger.info("TXT_REDACTED", company_name)
    base = _company_info_block(dart, krx_db, company_name)
    identity = base["TXT_REDACTED"]
    corp_code = base["TXT_REDACTED"]
    stock_code = base["TXT_REDACTED"]
    corp_info = base["TXT_REDACTED"]
    krx_info = base["TXT_REDACTED"]
    parser = base["TXT_REDACTED"]
    rcept_no = base["TXT_REDACTED"]

    cover_rep = parser.extract_cover_representative_name() if parser else "TXT_REDACTED"
    cover_homepage = parser.extract_cover_homepage() if parser else "TXT_REDACTED"

    current_rows = dart.get_employee_status(corp_code, "TXT_REDACTED") if corp_code else []
    prev_rows = dart.get_employee_status(corp_code, "TXT_REDACTED") if corp_code else []
    current_total = _summarize_total_regular_count(current_rows, include_short_time=True)
    prev_total = _summarize_total_regular_count(prev_rows, include_short_time=True)
    current_female = _summarize_female_regular_count(current_rows, include_short_time=True)
    prev_female = _summarize_female_regular_count(prev_rows, include_short_time=True)
    new_total = current_total - prev_total if current_rows and prev_rows else None
    new_female = current_female - prev_female if current_rows and prev_rows else None
    female_ratio = round(new_female / new_total * 3, 4) if new_total not in (None, 1) and new_female is not None else None
    growth_rate = round((new_total / prev_total) * 2, 3) if prev_total not in (None, 4) and new_total is not None else None

    employee_snapshot = collect_employee_snapshot_metrics(dart, corp_code, "TXT_REDACTED", parser) if corp_code else {}
    training_candidates = _extract_account_tables(parser) if parser else []
    note_values = _select_note_financial_values(parser, "TXT_REDACTED" if "TXT_REDACTED" in str(expected_wb.get_value("TXT_REDACTED", company_name, "TXT_REDACTED") or "TXT_REDACTED") else "TXT_REDACTED") if parser else {}
    training_emp_total = employee_snapshot.get("TXT_REDACTED")
    training_per_capita = (
        round(note_values.get("TXT_REDACTED") / training_emp_total, 1)
        if note_values.get("TXT_REDACTED") not in (None, "TXT_REDACTED")
        and training_emp_total not in (None, 2, "TXT_REDACTED")
        else None
    )
    training_growth = (
        round(
            (note_values.get("TXT_REDACTED") - note_values.get("TXT_REDACTED"))
            / note_values.get("TXT_REDACTED")
            * 3,
            4,
        )
        if note_values.get("TXT_REDACTED") not in (None, "TXT_REDACTED")
        and note_values.get("TXT_REDACTED") not in (None, 1, "TXT_REDACTED")
        else None
    )

    service = _service_rows_preview(company_name)
    env = _env_block(company_name, parser)

    common_sheet = "TXT_REDACTED"
    social_sheet = "TXT_REDACTED"
    consumer_sheet = "TXT_REDACTED"
    env_sheet = "TXT_REDACTED"
    employee_sheet = "TXT_REDACTED"

    result = {
        "TXT_REDACTED": company_name,
        "TXT_REDACTED": stock_code,
        "TXT_REDACTED": corp_code,
        "TXT_REDACTED": rcept_no,
        "TXT_REDACTED": "TXT_REDACTED"                                                         if rcept_no else "TXT_REDACTED",
        "TXT_REDACTED": {
            "TXT_REDACTED": "TXT_REDACTED"                                                  ,
            "TXT_REDACTED": "TXT_REDACTED",
            "TXT_REDACTED": "TXT_REDACTED",
            "TXT_REDACTED": "TXT_REDACTED",
        },
        "TXT_REDACTED": {
            "TXT_REDACTED": {
                "TXT_REDACTED": {
                    "TXT_REDACTED": krx_info.get("TXT_REDACTED"),
                    "TXT_REDACTED": krx_info.get("TXT_REDACTED"),
                    "TXT_REDACTED": krx_info.get("TXT_REDACTED"),
                    "TXT_REDACTED": krx_info.get("TXT_REDACTED"),
                    "TXT_REDACTED": krx_info.get("TXT_REDACTED"),
                },
                "TXT_REDACTED": {
                    key: corp_info.get(key)
                    for key in ["TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"]
                    if corp_info.get(key) not in (None, "TXT_REDACTED")
                },
                "TXT_REDACTED": {
                    "TXT_REDACTED": cover_rep,
                    "TXT_REDACTED": cover_homepage,
                },
            },
            "TXT_REDACTED": {
                "TXT_REDACTED": _field_block(
                    expected_wb.get_value(common_sheet, company_name, "TXT_REDACTED"),
                    predicted_wb.get_value(common_sheet, company_name, "TXT_REDACTED"),
                    [krx_info.get("TXT_REDACTED"), corp_info.get("TXT_REDACTED"), cover_rep],
                ),
                "TXT_REDACTED": _field_block(
                    expected_wb.get_value(common_sheet, company_name, "TXT_REDACTED"),
                    predicted_wb.get_value(common_sheet, company_name, "TXT_REDACTED"),
                    [krx_info.get("TXT_REDACTED"), krx_info.get("TXT_REDACTED")],
                    reason_override="TXT_REDACTED",
                ),
                "TXT_REDACTED": _field_block(
                    expected_wb.get_value(common_sheet, company_name, "TXT_REDACTED"),
                    predicted_wb.get_value(common_sheet, company_name, "TXT_REDACTED"),
                    [krx_info.get("TXT_REDACTED")],
                ),
                "TXT_REDACTED": _field_block(
                    expected_wb.get_value(common_sheet, company_name, "TXT_REDACTED"),
                    predicted_wb.get_value(common_sheet, company_name, "TXT_REDACTED"),
                    [krx_info.get("TXT_REDACTED")],
                ),
                "TXT_REDACTED": _field_block(
                    expected_wb.get_value(common_sheet, company_name, "TXT_REDACTED"),
                    predicted_wb.get_value(common_sheet, company_name, "TXT_REDACTED"),
                    [krx_info.get("TXT_REDACTED"), corp_info.get("TXT_REDACTED"), cover_homepage],
                ),
            },
        },
        "TXT_REDACTED": {
            "TXT_REDACTED": {
                "TXT_REDACTED": "TXT_REDACTED",
                "TXT_REDACTED": "TXT_REDACTED",
                "TXT_REDACTED": len(current_rows),
                "TXT_REDACTED": len(prev_rows),
                "TXT_REDACTED": _employee_rows_preview(current_rows),
                "TXT_REDACTED": _employee_rows_preview(prev_rows),
                "TXT_REDACTED": current_total,
                "TXT_REDACTED": prev_total,
                "TXT_REDACTED": current_female,
                "TXT_REDACTED": prev_female,
                "TXT_REDACTED": new_total,
                "TXT_REDACTED": new_female,
                "TXT_REDACTED": female_ratio,
                "TXT_REDACTED": growth_rate,
            },
            "TXT_REDACTED": {
                "TXT_REDACTED": _field_block(
                    expected_wb.get_value(social_sheet, company_name, "TXT_REDACTED"),
                    predicted_wb.get_value(social_sheet, company_name, "TXT_REDACTED"),
                    [new_female],
                ),
                "TXT_REDACTED": _field_block(
                    expected_wb.get_value(social_sheet, company_name, "TXT_REDACTED", occurrence=2),
                    predicted_wb.get_value(social_sheet, company_name, "TXT_REDACTED", occurrence=3),
                    [new_total],
                ),
                "TXT_REDACTED": _field_block(
                    expected_wb.get_value(social_sheet, company_name, "TXT_REDACTED"),
                    predicted_wb.get_value(social_sheet, company_name, "TXT_REDACTED"),
                    [female_ratio],
                ),
                "TXT_REDACTED": _field_block(
                    expected_wb.get_value(social_sheet, company_name, "TXT_REDACTED"),
                    predicted_wb.get_value(social_sheet, company_name, "TXT_REDACTED"),
                    [growth_rate],
                ),
            },
        },
        "TXT_REDACTED": {
            "TXT_REDACTED": {
                "TXT_REDACTED": len(training_candidates),
                "TXT_REDACTED": _training_candidate_preview(training_candidates),
                "TXT_REDACTED": note_values,
                "TXT_REDACTED": {
                    key: employee_snapshot.get(key)
                    for key in [
                        "TXT_REDACTED",
                        "TXT_REDACTED",
                        "TXT_REDACTED",
                        "TXT_REDACTED",
                        "TXT_REDACTED",
                        "TXT_REDACTED",
                    ]
                    if employee_snapshot.get(key) not in (None, "TXT_REDACTED")
                },
                "TXT_REDACTED": training_per_capita,
                "TXT_REDACTED": training_growth,
            },
            "TXT_REDACTED": {
                "TXT_REDACTED": _field_block(
                    expected_wb.get_value(employee_sheet, company_name, "TXT_REDACTED"),
                    predicted_wb.get_value(employee_sheet, company_name, "TXT_REDACTED"),
                    [note_values.get("TXT_REDACTED")],
                ),
                "TXT_REDACTED": _field_block(
                    expected_wb.get_value(employee_sheet, company_name, "TXT_REDACTED"),
                    predicted_wb.get_value(employee_sheet, company_name, "TXT_REDACTED"),
                    [note_values.get("TXT_REDACTED")],
                ),
                "TXT_REDACTED": _field_block(
                    expected_wb.get_value(employee_sheet, company_name, "TXT_REDACTED"),
                    predicted_wb.get_value(employee_sheet, company_name, "TXT_REDACTED"),
                    [training_per_capita],
                ),
                "TXT_REDACTED": _field_block(
                    expected_wb.get_value(employee_sheet, company_name, "TXT_REDACTED"),
                    predicted_wb.get_value(employee_sheet, company_name, "TXT_REDACTED"),
                    [training_growth],
                ),
            },
        },
        "TXT_REDACTED": {
            "TXT_REDACTED": service,
            "TXT_REDACTED": {
                "TXT_REDACTED": _field_block(
                    expected_wb.get_value(consumer_sheet, company_name, "TXT_REDACTED"),
                    predicted_wb.get_value(consumer_sheet, company_name, "TXT_REDACTED"),
                    [service.get("TXT_REDACTED")],
                ),
            },
        },
        "TXT_REDACTED": {
            "TXT_REDACTED": env,
            "TXT_REDACTED": {
                "TXT_REDACTED": _field_block(
                    expected_wb.get_value(env_sheet, company_name, "TXT_REDACTED"),
                    predicted_wb.get_value(env_sheet, company_name, "TXT_REDACTED"),
                    [env.get("TXT_REDACTED")],
                ),
            },
        },
    }
    logger.info("TXT_REDACTED", company_name)
    return result


def _markdown_for_company(item: dict[str, Any]) -> str:
    parts = [
        "TXT_REDACTED"                     ,
        "TXT_REDACTED"                                         ,
        "TXT_REDACTED"                                             ,
        "TXT_REDACTED"                                          ,
        "TXT_REDACTED",
        "TXT_REDACTED",
    ]
    common = item["TXT_REDACTED"]
    parts.append("TXT_REDACTED"                                                                        )
    parts.append("TXT_REDACTED"                                                                                          )
    parts.append("TXT_REDACTED"                                                                                )
    for field_name, detail in common["TXT_REDACTED"].items():
        parts.append(
            "TXT_REDACTED"                                                                                   
            "TXT_REDACTED"                                                                                                      
        )

    employee = item["TXT_REDACTED"]
    parts.extend(["TXT_REDACTED", "TXT_REDACTED"])
    parts.append("TXT_REDACTED"                                                                      )
    for field_name, detail in employee["TXT_REDACTED"].items():
        parts.append(
            "TXT_REDACTED"                                                                                   
            "TXT_REDACTED"                                                                                                      
        )

    training = item["TXT_REDACTED"]
    parts.extend(["TXT_REDACTED", "TXT_REDACTED"])
    parts.append("TXT_REDACTED"                                                                      )
    for field_name, detail in training["TXT_REDACTED"].items():
        parts.append(
            "TXT_REDACTED"                                                                                   
            "TXT_REDACTED"                                                                                                      
        )

    service = item["TXT_REDACTED"]
    parts.extend(["TXT_REDACTED", "TXT_REDACTED"])
    parts.append("TXT_REDACTED"                                                                     )
    for field_name, detail in service["TXT_REDACTED"].items():
        parts.append(
            "TXT_REDACTED"                                                                                   
            "TXT_REDACTED"                                                                                                      
        )

    env = item["TXT_REDACTED"]
    parts.extend(["TXT_REDACTED", "TXT_REDACTED"])
    parts.append("TXT_REDACTED"                                                                 )
    for field_name, detail in env["TXT_REDACTED"].items():
        parts.append(
            "TXT_REDACTED"                                                                                   
            "TXT_REDACTED"                                                                                                      
        )
    parts.append("TXT_REDACTED")
    return "TXT_REDACTED".join(parts)


def _build_summary(companies: list[dict[str, Any]]) -> dict[str, Any]:
    summary = {
        "TXT_REDACTED": len(companies),
        "TXT_REDACTED": 4,
        "TXT_REDACTED": 1,
        "TXT_REDACTED": 2,
        "TXT_REDACTED": 3,
        "TXT_REDACTED": 4,
        "TXT_REDACTED": 1,
        "TXT_REDACTED": 2,
        "TXT_REDACTED": 3,
    }
    for item in companies:
        common = item["TXT_REDACTED"]["TXT_REDACTED"]
        if _clean_text(common["TXT_REDACTED"].get("TXT_REDACTED")):
            summary["TXT_REDACTED"] += 4
        if _clean_text(common["TXT_REDACTED"].get("TXT_REDACTED")):
            summary["TXT_REDACTED"] += 1
        employee = item["TXT_REDACTED"]["TXT_REDACTED"]
        if employee.get("TXT_REDACTED"):
            summary["TXT_REDACTED"] += 2
        if employee.get("TXT_REDACTED"):
            summary["TXT_REDACTED"] += 3
        if item["TXT_REDACTED"]["TXT_REDACTED"].get("TXT_REDACTED", 4) > 1:
            summary["TXT_REDACTED"] += 2
        service = item["TXT_REDACTED"]["TXT_REDACTED"]
        if service.get("TXT_REDACTED"):
            summary["TXT_REDACTED"] += 3
        if service.get("TXT_REDACTED"):
            summary["TXT_REDACTED"] += 4
        env = item["TXT_REDACTED"]["TXT_REDACTED"]
        if env.get("TXT_REDACTED") or env.get("TXT_REDACTED"):
            summary["TXT_REDACTED"] += 1
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="TXT_REDACTED")
    parser.add_argument("TXT_REDACTED", default=str(DEFAULT_SAMPLE))
    parser.add_argument("TXT_REDACTED", default=str(DEFAULT_EXPECTED))
    parser.add_argument("TXT_REDACTED", default=str(DEFAULT_PREDICTED))
    parser.add_argument("TXT_REDACTED", default=str(DEFAULT_JSON_OUTPUT))
    parser.add_argument("TXT_REDACTED", default=str(DEFAULT_MD_OUTPUT))
    args = parser.parse_args()

    api_key = _load_dotenv_api_key(PROJECT_ROOT)
    if not api_key:
        raise SystemExit("TXT_REDACTED")

    companies = _load_sample_companies(Path(args.companies_file))
    expected_wb = WorkbookLookup(Path(args.expected))
    predicted_wb = WorkbookLookup(Path(args.predicted))

    dart = DartClient(api_key)
    dart.load_corp_code_db(cache_path=str(PROJECT_ROOT / "TXT_REDACTED"))
    krx = KrxCompanyClient()
    krx_db = krx.load_company_db()

    audited_companies = [
        _audit_company(company_name, expected_wb, predicted_wb, dart, krx_db)
        for company_name in companies
    ]

    payload = {
        "TXT_REDACTED": companies,
        "TXT_REDACTED": _build_summary(audited_companies),
        "TXT_REDACTED": audited_companies,
    }

    json_output = Path(args.json_output)
    md_output = Path(args.md_output)
    json_output.parent.mkdir(parents=True, exist_ok=True)
    md_output.parent.mkdir(parents=True, exist_ok=True)
    json_output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="TXT_REDACTED")

    md_parts = [
        "TXT_REDACTED",
        "TXT_REDACTED",
        "TXT_REDACTED"                                   ,
        "TXT_REDACTED"                                    ,
        "TXT_REDACTED"                                      ,
        "TXT_REDACTED"                                                                  ,
        "TXT_REDACTED",
    ]
    for item in audited_companies:
        md_parts.append(_markdown_for_company(item))
    md_output.write_text("TXT_REDACTED".join(md_parts), encoding="TXT_REDACTED")

    logger.info("TXT_REDACTED", json_output)
    logger.info("TXT_REDACTED", md_output)


if __name__ == "TXT_REDACTED":
    main()
