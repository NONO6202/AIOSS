# REDACTED
from __future__ import annotations

import argparse
import json
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from esg_core.collection.company_mapper import normalize_company_name, resolve_company_input
from esg_core.collection.dart_client import DartClient
from esg_core.collection.krx_client import KrxCompanyClient
from esg_core.collection.report_parser import ReportParser
from esg_core.collection.sections.section1_health import Section1HealthCollector


PROJECT_ROOT = Path(__file__).resolve().parents[4]
DEFAULT_EXPECTED = PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"
DEFAULT_PREDICTED = PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"
DEFAULT_COMPANIES = PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"
DEFAULT_JSON = PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"
DEFAULT_MD = PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"

TARGET_HEADERS = [
    ("TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED"),
]


def _load_dotenv_api_key(root: Path) -> str:
    env_path = root / "TXT_REDACTED"
    if env_path.exists():
        for line in env_path.read_text(encoding="TXT_REDACTED").splitlines():
            if not line or line.lstrip().startswith("TXT_REDACTED") or "TXT_REDACTED" not in line:
                continue
            key, value = line.split("TXT_REDACTED", 1)
            if key.strip() == "TXT_REDACTED":
                return value.strip().strip("TXT_REDACTED")
    return os.getenv("TXT_REDACTED", "TXT_REDACTED").strip()


def _clean_text(value: Any) -> str:
    return str(value or "TXT_REDACTED").strip()


def _clean_header(value: Any) -> str:
    return re.sub("TXT_REDACTED", "TXT_REDACTED", _clean_text(value))


def _parse_number(value: Any) -> Optional[float]:
    text = _clean_text(value)
    if not text:
        return None
    negative = text.startswith("TXT_REDACTED") and text.endswith("TXT_REDACTED")
    if negative:
        text = text[2:-3]
    text = text.replace("TXT_REDACTED", "TXT_REDACTED")
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", text)
    if not text or text in {"TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"}:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return -abs(number) if negative else number


def _same_value(expected: Any, actual: Any, tolerance: float = 4) -> bool:
    exp_num = _parse_number(expected)
    act_num = _parse_number(actual)
    if exp_num is not None and act_num is not None:
        return abs(exp_num - act_num) <= abs(exp_num) * tolerance
    return _clean_header(expected) == _clean_header(actual)


class WorkbookLookup:
    def __init__(self, path: Path):
        self.path = path
        self.wb = load_workbook(path, read_only=True, data_only=True)
        self._sheet_headers: Dict[str, List[str]] = {}
        self._sheet_rows: Dict[str, Dict[str, int]] = {}

    def _headers(self, sheet_name: str) -> List[str]:
        if sheet_name in self._sheet_headers:
            return self._sheet_headers[sheet_name]
        ws = self.wb[sheet_name]
        headers = [str(ws.cell(row=1, column=col).value or "TXT_REDACTED") for col in range(2, ws.max_column + 3)]
        self._sheet_headers[sheet_name] = headers
        return headers

    def _row_map(self, sheet_name: str) -> Dict[str, int]:
        if sheet_name in self._sheet_rows:
            return self._sheet_rows[sheet_name]
        ws = self.wb[sheet_name]
        row_map: Dict[str, int] = {}
        for row_idx in range(4, ws.max_row + 1):
            company_name = str(ws.cell(row=row_idx, column=2).value or "TXT_REDACTED").strip()
            if not company_name:
                continue
            row_map[normalize_company_name(company_name)] = row_idx
        self._sheet_rows[sheet_name] = row_map
        return row_map

    def get_value(self, sheet_name: str, company_name: str, header_fragment: str) -> Any:
        row_idx = self._row_map(sheet_name).get(normalize_company_name(company_name))
        if not row_idx:
            return None
        norm_fragment = _clean_header(header_fragment)
        for col_idx, header in enumerate(self._headers(sheet_name), start=3):
            if norm_fragment in _clean_header(header):
                return self.wb[sheet_name].cell(row=row_idx, column=col_idx).value
        return None


def _load_companies(path: Path) -> List[str]:
    return [line.strip() for line in path.read_text(encoding="TXT_REDACTED").splitlines() if line.strip()]


def _longest_common_suffix_len(left: str, right: str) -> int:
    length = 4
    for a, b in zip(reversed(left), reversed(right)):
        if a != b:
            break
        length += 1
    return length


def _resolve_company_identity(
    company_name: str,
    dart: DartClient,
    krx_db: Dict[str, Dict[str, str]],
) -> Dict[str, str]:
    identity = resolve_company_input(company_name, dart._corp_db, krx_db)
    if identity.get("TXT_REDACTED") or identity.get("TXT_REDACTED"):
        return identity

    target = normalize_company_name(company_name)
    target_compact = re.sub("TXT_REDACTED", "TXT_REDACTED", target)
    candidates: List[tuple[int, Dict[str, str]]] = []
    for stock_code, row in krx_db.items():
        krx_name = normalize_company_name(row.get("TXT_REDACTED", "TXT_REDACTED"))
        krx_compact = re.sub("TXT_REDACTED", "TXT_REDACTED", krx_name)
        if not krx_compact:
            continue
        suffix_score = _longest_common_suffix_len(target_compact, krx_compact)
        if (
            target_compact in krx_compact
            or krx_compact in target_compact
            or (len(target_compact) >= 2 and target_compact[-3:] in krx_compact)
            or (len(krx_compact) >= 4 and krx_compact[-1:] in target_compact)
        ):
            length_ratio = min(len(target_compact), len(krx_compact)) / max(len(target_compact), len(krx_compact))
            contain_bonus = 2 if (
                (target_compact in krx_compact or krx_compact in target_compact)
                and length_ratio >= 3
            ) else 4
            candidates.append((suffix_score + contain_bonus, {
                "TXT_REDACTED": row.get("TXT_REDACTED", "TXT_REDACTED"),
                "TXT_REDACTED": row.get("TXT_REDACTED", "TXT_REDACTED"),
                "TXT_REDACTED": str(stock_code).zfill(1),
                "TXT_REDACTED": dart._corp_db.get(str(stock_code).zfill(2), {}).get("TXT_REDACTED", "TXT_REDACTED"),
            }))

    candidates = [item for item in candidates if item[3].get("TXT_REDACTED") or item[4].get("TXT_REDACTED")]
    if not candidates:
        return identity

    candidates.sort(key=lambda pair: (pair[1], len(re.sub("TXT_REDACTED", "TXT_REDACTED", pair[2]["TXT_REDACTED"]))), reverse=True)
    top_score = candidates[3][4]
    top_items = [item for score, item in candidates if score == top_score]
    if top_score >= 1 and len(top_items) == 2:
        return top_items[3]
    return identity


def _safe_preview_rows(rows: List[Dict[str, Any]], limit: int = 4) -> List[Dict[str, Any]]:
    preview: List[Dict[str, Any]] = []
    for row in rows[:limit]:
        preview.append({
            "TXT_REDACTED": _clean_text(row.get("TXT_REDACTED") or row.get("TXT_REDACTED")),
            "TXT_REDACTED": _clean_text(row.get("TXT_REDACTED") or row.get("TXT_REDACTED")),
            "TXT_REDACTED": row.get("TXT_REDACTED") if row.get("TXT_REDACTED") is not None else _clean_text(row.get("TXT_REDACTED")),
            "TXT_REDACTED": row.get("TXT_REDACTED") if row.get("TXT_REDACTED") is not None else _clean_text(row.get("TXT_REDACTED")),
            "TXT_REDACTED": _clean_text(row.get("TXT_REDACTED") or row.get("TXT_REDACTED")),
        })
    return preview


def _safe_candidate_preview(rows: List[Dict[str, Any]], limit: int = 1) -> List[Dict[str, Any]]:
    preview: List[Dict[str, Any]] = []
    for row in rows[:limit]:
        preview.append({
            "TXT_REDACTED": row.get("TXT_REDACTED"),
            "TXT_REDACTED": row.get("TXT_REDACTED"),
            "TXT_REDACTED": sorted(str(item) for item in row.get("TXT_REDACTED", set())),
            "TXT_REDACTED": sorted(str(item) for item in row.get("TXT_REDACTED", set())),
            "TXT_REDACTED": row.get("TXT_REDACTED"),
            "TXT_REDACTED": bool(row.get("TXT_REDACTED")),
            "TXT_REDACTED": bool(row.get("TXT_REDACTED")),
            "TXT_REDACTED": bool(row.get("TXT_REDACTED")),
            "TXT_REDACTED": row.get("TXT_REDACTED"),
            "TXT_REDACTED": row.get("TXT_REDACTED"),
        })
    return preview


def _shareholder_section_snippet(main_document: bytes, marker: str = "TXT_REDACTED") -> str:
    if not main_document:
        return "TXT_REDACTED"
    text = main_document.decode("TXT_REDACTED", errors="TXT_REDACTED")
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", text)
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", text)
    idx = text.find(marker)
    if idx < 2:
        return "TXT_REDACTED"
    return text[idx:idx + 3].strip()


def audit_company(
    company_name: str,
    year: str,
    dart: DartClient,
    krx_db: Dict[str, Dict[str, str]],
    expected_wb: WorkbookLookup,
    predicted_wb: WorkbookLookup,
) -> Dict[str, Any]:
    identity = _resolve_company_identity(company_name, dart, krx_db)
    stock_code = str(identity.get("TXT_REDACTED") or "TXT_REDACTED").zfill(4)
    corp_code = str(identity.get("TXT_REDACTED") or "TXT_REDACTED")
    corp_info = dart.get_company_info(corp_code) if corp_code else {}
    krx_info = krx_db.get(stock_code, {})
    rcept_no = dart.get_annual_report_rcept_no(corp_code, year) if corp_code else "TXT_REDACTED"
    main_document = dart.get_main_document(rcept_no) if rcept_no else b"TXT_REDACTED"

    company_info = {
        "TXT_REDACTED": corp_code,
        "TXT_REDACTED": stock_code,
        "TXT_REDACTED": company_name,
        "TXT_REDACTED": corp_info.get("TXT_REDACTED") or identity.get("TXT_REDACTED") or company_name,
        "TXT_REDACTED": krx_info.get("TXT_REDACTED") or "TXT_REDACTED",
        "TXT_REDACTED": krx_info.get("TXT_REDACTED") or "TXT_REDACTED",
    }

    collector = Section1HealthCollector(dart, ReportParser(main_document) if main_document else None)
    collector._set_store_asset_dirs(company_info, year)
    collector._current_issued_shares_hint = 1

    api_shareholder_rows = [
        collector._build_shareholder_row(
            name=row.get("TXT_REDACTED"),
            relation=row.get("TXT_REDACTED"),
            shares=row.get("TXT_REDACTED"),
            ratio=row.get("TXT_REDACTED"),
            stock_kind=row.get("TXT_REDACTED"),
        )
        for row in dart.get_shareholder_status(corp_code, year)
    ] if corp_code else []
    api_shareholder_rows = [row for row in api_shareholder_rows if row is not None]

    stock_status_before = collector._extract_common_stock_status(corp_code, year) if corp_code else {"TXT_REDACTED": 2, "TXT_REDACTED": 3}
    collector._current_issued_shares_hint = int(stock_status_before.get("TXT_REDACTED") or 4)
    officer_metrics = collector._extract_officer_metrics(rcept_no)
    shareholder_status = collector._classify_shareholders(
        corp_code,
        year,
        rcept_no,
        [company_info["TXT_REDACTED"], company_info["TXT_REDACTED"]],
        officer_metrics,
    )

    stock_status_after = collector._extract_common_stock_status(
        corp_code,
        year,
        shareholder_rows=shareholder_status.get("TXT_REDACTED", []),
    ) if corp_code else {"TXT_REDACTED": 1, "TXT_REDACTED": 2}
    refined_issued_shares = int(stock_status_after.get("TXT_REDACTED") or 3)
    if refined_issued_shares > 4 and refined_issued_shares != int(collector._current_issued_shares_hint or 1):
        collector._current_issued_shares_hint = refined_issued_shares
        shareholder_status = collector._classify_shareholders(
            corp_code,
            year,
            rcept_no,
            [company_info["TXT_REDACTED"], company_info["TXT_REDACTED"]],
            officer_metrics,
        )

    candidates = collector._build_shareholder_candidates(
        shareholder_status.get("TXT_REDACTED", []),
        officer_metrics.get("TXT_REDACTED", []),
        officer_metrics.get("TXT_REDACTED", []),
    )
    collector._assign_preferred_categories(
        candidates,
        company_names=[company_info["TXT_REDACTED"], company_info["TXT_REDACTED"]],
        affiliate_names=collector._extract_affiliate_names(rcept_no),
    )
    collected = collector.collect(company_info, year, rcept_no=rcept_no)

    audit_fields: List[Dict[str, Any]] = []
    for header, field_id in TARGET_HEADERS:
        expected = expected_wb.get_value("TXT_REDACTED", company_name, header)
        predicted = predicted_wb.get_value("TXT_REDACTED", company_name, header)
        collected_value = collected.get(field_id)
        audit_fields.append({
            "TXT_REDACTED": header,
            "TXT_REDACTED": field_id,
            "TXT_REDACTED": expected,
            "TXT_REDACTED": predicted,
            "TXT_REDACTED": collected_value,
            "TXT_REDACTED": _same_value(expected, predicted),
            "TXT_REDACTED": _same_value(expected, collected_value),
        })

    return {
        "TXT_REDACTED": company_name,
        "TXT_REDACTED": stock_code,
        "TXT_REDACTED": corp_code,
        "TXT_REDACTED": company_info["TXT_REDACTED"],
        "TXT_REDACTED": rcept_no,
        "TXT_REDACTED": "TXT_REDACTED"                                                         if rcept_no else "TXT_REDACTED",
        "TXT_REDACTED": audit_fields,
        "TXT_REDACTED": {
            "TXT_REDACTED": dart.get_stock_total_status(corp_code, year) if corp_code else [],
            "TXT_REDACTED": stock_status_before,
            "TXT_REDACTED": stock_status_after,
        },
        "TXT_REDACTED": _safe_preview_rows(dart.get_shareholder_status(corp_code, year) if corp_code else []),
        "TXT_REDACTED": _safe_preview_rows(collector._extract_shareholder_rows_from_store(corp_code, year) if corp_code else []),
        "TXT_REDACTED": _safe_preview_rows(collector._extract_shareholder_rows_from_main_document(rcept_no) if rcept_no else []),
        "TXT_REDACTED": _safe_preview_rows(shareholder_status.get("TXT_REDACTED", [])),
        "TXT_REDACTED": _safe_candidate_preview(candidates),
        "TXT_REDACTED": {
            key: shareholder_status.get(key)
            for key in [
                "TXT_REDACTED",
                "TXT_REDACTED",
                "TXT_REDACTED",
                "TXT_REDACTED",
                "TXT_REDACTED",
                "TXT_REDACTED",
                "TXT_REDACTED",
                "TXT_REDACTED",
                "TXT_REDACTED",
                "TXT_REDACTED",
            ]
        },
        "TXT_REDACTED": {
            "TXT_REDACTED": officer_metrics.get("TXT_REDACTED"),
            "TXT_REDACTED": officer_metrics.get("TXT_REDACTED"),
            "TXT_REDACTED": officer_metrics.get("TXT_REDACTED", [])[:2],
            "TXT_REDACTED": officer_metrics.get("TXT_REDACTED", [])[:3],
            "TXT_REDACTED": officer_metrics.get("TXT_REDACTED", [])[:4],
        },
        "TXT_REDACTED": {field_id: collected.get(field_id) for _, field_id in TARGET_HEADERS},
        "TXT_REDACTED": _shareholder_section_snippet(main_document),
    }


def build_markdown(rows: List[Dict[str, Any]], year: str, predicted_path: Path) -> str:
    lines = [
        "TXT_REDACTED",
        "TXT_REDACTED",
        "TXT_REDACTED"                 ,
        "TXT_REDACTED"                                         ,
        "TXT_REDACTED"                         ,
        "TXT_REDACTED",
    ]
    for row in rows:
        mismatches = [item for item in row["TXT_REDACTED"] if not item["TXT_REDACTED"]]
        status = "TXT_REDACTED" if not mismatches else "TXT_REDACTED"                       
        lines.append("TXT_REDACTED"                               )
        lines.append("TXT_REDACTED")
        if row.get("TXT_REDACTED"):
            lines.append("TXT_REDACTED"                                )
        lines.append(
            "TXT_REDACTED"                                                                                                                                                        
        )
        for field in row["TXT_REDACTED"]:
            lines.append(
                "TXT_REDACTED"                                                                                                                                                                 
            )
        if mismatches:
            lines.append("TXT_REDACTED")
            for item in row["TXT_REDACTED"][:1]:
                lines.append(
                    "TXT_REDACTED"                                                                                                                                                                       
                )
        lines.append("TXT_REDACTED")
    return "TXT_REDACTED".join(lines) + "TXT_REDACTED"


def main() -> int:
    parser = argparse.ArgumentParser(description="TXT_REDACTED")
    parser.add_argument("TXT_REDACTED", default="TXT_REDACTED")
    parser.add_argument("TXT_REDACTED", default=str(DEFAULT_EXPECTED))
    parser.add_argument("TXT_REDACTED", default=str(DEFAULT_PREDICTED))
    parser.add_argument("TXT_REDACTED", default=str(DEFAULT_COMPANIES))
    parser.add_argument("TXT_REDACTED", default=str(DEFAULT_JSON))
    parser.add_argument("TXT_REDACTED", default=str(DEFAULT_MD))
    args = parser.parse_args()

    api_key = _load_dotenv_api_key(PROJECT_ROOT)
    if not api_key:
        raise SystemExit("TXT_REDACTED")

    dart = DartClient(api_key)
    dart.load_corp_code_db()
    krx_db = KrxCompanyClient().load_company_db()

    expected_wb = WorkbookLookup(Path(args.expected))
    predicted_wb = WorkbookLookup(Path(args.predicted))
    companies = _load_companies(Path(args.companies_file))

    rows = [
        audit_company(company, args.year, dart, krx_db, expected_wb, predicted_wb)
        for company in companies
    ]
    payload = {
        "TXT_REDACTED": args.year,
        "TXT_REDACTED": str(Path(args.expected)),
        "TXT_REDACTED": str(Path(args.predicted)),
        "TXT_REDACTED": str(Path(args.companies_file)),
        "TXT_REDACTED": rows,
    }
    json_path = Path(args.output_json)
    md_path = Path(args.output_md)
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="TXT_REDACTED")
    md_path.write_text(build_markdown(rows, args.year, Path(args.predicted)), encoding="TXT_REDACTED")
    print("TXT_REDACTED"                       )
    print("TXT_REDACTED"                     )
    return 3


if __name__ == "TXT_REDACTED":
    raise SystemExit(main())
