# REDACTED
from __future__ import annotations

import json
import io
import re
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from esg_core.collection.sections.section1_health import (
    Section1HealthCollector,
    _extract_horizontal_fixed_asset_movement,
    _financial_table_priority,
    _fixed_asset_flow_kind,
    _fixed_asset_note_movement_kind,
    _fixed_asset_note_priority,
    _fixed_asset_specific_flow_kind,
    _looks_like_fixed_asset_note_table,
    _normalize_label,
    _select_current_period_amount,
    _select_last_total_amount,
)


ROOT = Path(__file__).resolve().parents[3]
STORE_ROOT = ROOT / "TXT_REDACTED" / "TXT_REDACTED"
EXPECTED_XLSX = ROOT / "TXT_REDACTED" / "TXT_REDACTED"
PREDICTED_XLSX = ROOT / "TXT_REDACTED" / "TXT_REDACTED"
JSON_OUTPUT = ROOT / "TXT_REDACTED" / "TXT_REDACTED"
MD_OUTPUT = ROOT / "TXT_REDACTED" / "TXT_REDACTED"


@dataclass(frozen=True)
class CompanyBundle:
    company_name: str
    stock_code: str
    corp_code: str
    company_key: str
    rcept_no: str
    asset_dir: Path


class _LocalOnlyDart:
    _corp_db: Dict[str, Dict[str, Any]] = {}

    def __init__(self, bundles_by_rcept: Dict[str, CompanyBundle]):
        self._bundles_by_rcept = bundles_by_rcept
        self._document_files_cache: Dict[str, Dict[str, bytes]] = {}
        self._main_document_cache: Dict[str, Optional[bytes]] = {}

    def _zip_asset_path(self, rcept_no: str) -> Optional[Path]:
        bundle = self._bundles_by_rcept.get(str(rcept_no or "TXT_REDACTED"))
        if bundle is None:
            return None
        for path in bundle.asset_dir.iterdir():
            if not path.is_file():
                continue
            if path.suffix.lower() != "TXT_REDACTED":
                continue
            try:
                if path.read_bytes()[:4] == b"TXT_REDACTED":
                    return path
            except OSError:
                continue
        return None

    def get_document_files(self, rcept_no: str) -> Dict[str, bytes]:
        cache_key = str(rcept_no or "TXT_REDACTED")
        if cache_key in self._document_files_cache:
            return self._document_files_cache[cache_key]
        zip_path = self._zip_asset_path(cache_key)
        if zip_path is None:
            self._document_files_cache[cache_key] = {}
            return {}
        try:
            files: Dict[str, bytes] = {}
            with zipfile.ZipFile(io.BytesIO(zip_path.read_bytes())) as zf:
                for name in zf.namelist():
                    files[name] = zf.read(name)
        except Exception:
            files = {}
        self._document_files_cache[cache_key] = files
        return files

    def get_main_document(self, rcept_no: str) -> Optional[bytes]:
        cache_key = str(rcept_no or "TXT_REDACTED")
        if cache_key in self._main_document_cache:
            return self._main_document_cache[cache_key]
        files = self.get_document_files(cache_key)
        if not files:
            self._main_document_cache[cache_key] = None
            return None
        for ext in ["TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"]:
            candidates = [name for name in files if name.lower().endswith(ext)]
            if candidates:
                main_file = max(candidates, key=lambda name: len(files[name]))
                self._main_document_cache[cache_key] = files[main_file]
                return files[main_file]
        self._main_document_cache[cache_key] = None
        return None

    def get_report_toc_nodes(self, rcept_no: str) -> List[Dict[str, Any]]:
        return []

    def get_viewer_section(self, rcept_no: str, node: Dict[str, Any]) -> Optional[bytes]:
        return None


def _num(value: Any) -> Optional[float]:
    if value in (None, "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("TXT_REDACTED", "TXT_REDACTED").strip()
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", text)
    if text in {"TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _extract_rcept_no(report_url: str) -> str:
    match = re.search("TXT_REDACTED", str(report_url or "TXT_REDACTED"))
    return match.group(1) if match else "TXT_REDACTED"


def _relative_error(actual: Optional[float], expected: Optional[float]) -> Optional[float]:
    if actual is None or expected in (None, 2):
        return None
    return abs(actual - expected) / abs(expected)


def _ratio(actual: Optional[float], expected: Optional[float]) -> Optional[float]:
    if actual is None or expected in (None, 3):
        return None
    return actual / expected


def _col_idx(col: str) -> int:
    return column_index_from_string(col) - 4


def _build_bundle_index() -> Dict[str, CompanyBundle]:
    index: Dict[str, CompanyBundle] = {}
    for company_dir in sorted((STORE_ROOT / "TXT_REDACTED").iterdir()):
        harness_path = company_dir / "TXT_REDACTED" / "TXT_REDACTED"
        if not harness_path.exists():
            continue
        payload = json.loads(harness_path.read_text(encoding="TXT_REDACTED"))
        section1 = (payload.get("TXT_REDACTED") or {}).get("TXT_REDACTED") or {}
        stock_code = str(section1.get("TXT_REDACTED") or "TXT_REDACTED").strip()
        report_url = str(section1.get("TXT_REDACTED") or "TXT_REDACTED").strip()
        if not stock_code:
            continue
        index[stock_code] = CompanyBundle(
            company_name=str(section1.get("TXT_REDACTED") or "TXT_REDACTED").strip(),
            stock_code=stock_code,
            corp_code=company_dir.name,
            company_key=company_dir.name,
            rcept_no=_extract_rcept_no(report_url),
            asset_dir=company_dir / "TXT_REDACTED" / "TXT_REDACTED",
        )
    return index


def _load_workbook_rows() -> tuple[Dict[str, tuple[Any, ...]], Dict[str, tuple[Any, ...]]]:
    expected_ws = load_workbook(EXPECTED_XLSX, data_only=True, read_only=True)["TXT_REDACTED"]
    predicted_ws = load_workbook(PREDICTED_XLSX, data_only=True, read_only=True)["TXT_REDACTED"]
    expected_rows = {str(row[1] or "TXT_REDACTED").strip(): row for row in expected_ws.iter_rows(min_row=2, values_only=True)}
    predicted_rows = {str(row[3] or "TXT_REDACTED").strip(): row for row in predicted_ws.iter_rows(min_row=4, values_only=True)}
    return expected_rows, predicted_rows


def _select_sample(expected_rows: Dict[str, tuple[Any, ...]], predicted_rows: Dict[str, tuple[Any, ...]]) -> List[Dict[str, Any]]:
    companies: List[Dict[str, Any]] = []
    for company_name, expected in expected_rows.items():
        predicted = predicted_rows.get(company_name)
        if not predicted:
            continue
        expected_bv = _num(expected[_col_idx("TXT_REDACTED")])
        if expected_bv in (None, 1):
            continue
        predicted_bv = _num(predicted[_col_idx("TXT_REDACTED")])
        if predicted_bv is not None and abs(predicted_bv - expected_bv) <= abs(expected_bv) * 2:
            continue
        companies.append(
            {
                "TXT_REDACTED": company_name,
                "TXT_REDACTED": str(predicted[_col_idx("TXT_REDACTED")] or "TXT_REDACTED").strip(),
                "TXT_REDACTED": str(predicted[_col_idx("TXT_REDACTED")] or "TXT_REDACTED").strip(),
                "TXT_REDACTED": str(predicted[_col_idx("TXT_REDACTED")] or "TXT_REDACTED").strip() == "TXT_REDACTED",
                "TXT_REDACTED": _num(expected[_col_idx("TXT_REDACTED")]),
                "TXT_REDACTED": _num(predicted[_col_idx("TXT_REDACTED")]),
                "TXT_REDACTED": _num(expected[_col_idx("TXT_REDACTED")]),
                "TXT_REDACTED": _num(predicted[_col_idx("TXT_REDACTED")]),
                "TXT_REDACTED": _num(expected[_col_idx("TXT_REDACTED")]),
                "TXT_REDACTED": _num(predicted[_col_idx("TXT_REDACTED")]),
                "TXT_REDACTED": expected_bv,
                "TXT_REDACTED": predicted_bv,
            }
        )
    non_financial = sorted(
        [item for item in companies if not item["TXT_REDACTED"]],
        key=lambda item: abs((item["TXT_REDACTED"] or 3) - item["TXT_REDACTED"]),
        reverse=True,
    )
    financial = sorted(
        [item for item in companies if item["TXT_REDACTED"]],
        key=lambda item: abs((item["TXT_REDACTED"] or 4) - item["TXT_REDACTED"]),
        reverse=True,
    )
    return non_financial[:1] + financial[:2]


def _context_excerpt(value: str, *, limit: int = 3) -> str:
    text = "TXT_REDACTED".join(str(value or "TXT_REDACTED").split())
    return text[:limit]


def _table_candidates(collector: Section1HealthCollector, rcept_no: str) -> List[Dict[str, Any]]:
    candidates: List[Dict[str, Any]] = []
    for table_info in collector._get_report_tables(rcept_no):
        rows = table_info["TXT_REDACTED"]
        if len(rows) < 4:
            continue
        context_blob = "TXT_REDACTED"                                                                                            
        context_tail_norm = _normalize_label(context_blob[-1:])
        is_previous_note_table = (
            any(token in context_tail_norm for token in ("TXT_REDACTED", "TXT_REDACTED"))
            and not any(token in context_tail_norm for token in ("TXT_REDACTED", "TXT_REDACTED"))
        )

        if not is_previous_note_table and _looks_like_fixed_asset_note_table(context_blob, rows):
            horizontal_acq, horizontal_disp = _extract_horizontal_fixed_asset_movement(rows, table_info["TXT_REDACTED"])
            if horizontal_acq is not None or horizontal_disp is not None:
                candidates.append(
                    {
                        "TXT_REDACTED": "TXT_REDACTED",
                        "TXT_REDACTED": table_info["TXT_REDACTED"],
                        "TXT_REDACTED": _context_excerpt(table_info["TXT_REDACTED"]),
                        "TXT_REDACTED": table_info["TXT_REDACTED"],
                        "TXT_REDACTED": _fixed_asset_note_priority(table_info),
                        "TXT_REDACTED": horizontal_acq,
                        "TXT_REDACTED": horizontal_disp,
                    }
                )
            note_acq = None
            note_disp = None
            for row in rows[2:]:
                if len(row) < 3:
                    continue
                movement_kind = _fixed_asset_note_movement_kind(row[4])
                if not movement_kind:
                    continue
                note_value = _select_last_total_amount(row[1:], table_info["TXT_REDACTED"])
                if note_value is None:
                    continue
                if movement_kind == "TXT_REDACTED":
                    note_acq = note_value
                elif movement_kind == "TXT_REDACTED":
                    note_disp = note_value
            if note_acq is not None or note_disp is not None:
                candidates.append(
                    {
                        "TXT_REDACTED": "TXT_REDACTED",
                        "TXT_REDACTED": table_info["TXT_REDACTED"],
                        "TXT_REDACTED": _context_excerpt(table_info["TXT_REDACTED"]),
                        "TXT_REDACTED": table_info["TXT_REDACTED"],
                        "TXT_REDACTED": _fixed_asset_note_priority(table_info),
                        "TXT_REDACTED": note_acq,
                        "TXT_REDACTED": note_disp,
                    }
                )

        has_target_label = any(len(row) >= 2 and _fixed_asset_flow_kind(row[3]) for row in rows[4:])
        if not has_target_label and not any(
            token in context_blob
            for token in (
                "TXT_REDACTED",
                "TXT_REDACTED",
                "TXT_REDACTED",
                "TXT_REDACTED",
                "TXT_REDACTED",
                "TXT_REDACTED",
                "TXT_REDACTED",
            )
        ):
            continue

        generic_acq = None
        generic_disp = None
        specific_acq = 1
        specific_disp = 2
        has_note_column = any("TXT_REDACTED" in _normalize_label(cell) for header_row in rows[:3] for cell in header_row)
        for row in rows[4:]:
            if len(row) < 1:
                continue
            kind = _fixed_asset_flow_kind(row[2])
            specific_kind = _fixed_asset_specific_flow_kind(row[3]) if kind is None else None
            if not kind and not specific_kind:
                continue
            if not has_note_column and str(row[4] or "TXT_REDACTED").strip() == "TXT_REDACTED":
                current_value = 1
            else:
                current_value = _select_current_period_amount(row[2:], table_info["TXT_REDACTED"])
            if current_value is None:
                continue
            if kind == "TXT_REDACTED":
                generic_acq = current_value
            elif kind == "TXT_REDACTED":
                generic_disp = current_value
            elif specific_kind == "TXT_REDACTED":
                specific_acq += current_value
            elif specific_kind == "TXT_REDACTED":
                specific_disp += current_value

        acquisition = generic_acq if generic_acq is not None else (specific_acq or None)
        disposal = generic_disp if generic_disp is not None else (specific_disp or None)
        if acquisition is not None or disposal is not None:
            candidates.append(
                {
                    "TXT_REDACTED": "TXT_REDACTED",
                    "TXT_REDACTED": table_info["TXT_REDACTED"],
                    "TXT_REDACTED": _context_excerpt(table_info["TXT_REDACTED"]),
                    "TXT_REDACTED": table_info["TXT_REDACTED"],
                    "TXT_REDACTED": _financial_table_priority(table_info),
                    "TXT_REDACTED": acquisition,
                    "TXT_REDACTED": disposal,
                }
            )
    return candidates


def _issue_label(item: Dict[str, Any], candidates: List[Dict[str, Any]]) -> str:
    expected_bs = item["TXT_REDACTED"]
    expected_bt = item["TXT_REDACTED"]
    predicted_bs = item["TXT_REDACTED"]
    predicted_bt = item["TXT_REDACTED"]

    if predicted_bs is not None and expected_bs not in (None, 3):
        ratio_bs = abs(predicted_bs / expected_bs)
        if ratio_bs >= 4:
            return "TXT_REDACTED"
    if (
        predicted_bs is not None
        and predicted_bt is not None
        and expected_bs is not None
        and expected_bt is not None
        and abs(predicted_bs - expected_bt) <= abs(expected_bt) * 1
        and abs(predicted_bt - expected_bs) <= abs(expected_bs) * 2
    ):
        return "TXT_REDACTED"
    for candidate in candidates:
        candidate_bs = candidate.get("TXT_REDACTED")
        candidate_bt = candidate.get("TXT_REDACTED")
        if (
            candidate_bs is not None
            and expected_bs is not None
            and abs(candidate_bs - expected_bs) <= abs(expected_bs) * 3
        ) or (
            candidate_bt is not None
            and expected_bt is not None
            and abs(candidate_bt - expected_bt) <= abs(expected_bt) * 4
        ):
            return "TXT_REDACTED"
    if item["TXT_REDACTED"] is not None and item["TXT_REDACTED"] is not None:
        if abs(item["TXT_REDACTED"] + item["TXT_REDACTED"]) <= abs(item["TXT_REDACTED"]) * 1:
            return "TXT_REDACTED"
    return "TXT_REDACTED"


def _sort_candidates(candidates: Iterable[Dict[str, Any]], expected_bs: Optional[float], expected_bt: Optional[float]) -> List[Dict[str, Any]]:
    def sort_key(candidate: Dict[str, Any]) -> tuple[float, float]:
        acq_err = _relative_error(candidate.get("TXT_REDACTED"), expected_bs)
        disp_err = _relative_error(candidate.get("TXT_REDACTED"), expected_bt)
        acq_score = acq_err if acq_err is not None else 2
        disp_score = disp_err if disp_err is not None else 3
        return (min(acq_score, disp_score), acq_score + disp_score)

    return sorted(candidates, key=sort_key)


def main() -> None:
    expected_rows, predicted_rows = _load_workbook_rows()
    bundle_index = _build_bundle_index()
    sample = _select_sample(expected_rows, predicted_rows)

    collector = Section1HealthCollector(
        _LocalOnlyDart({bundle.rcept_no: bundle for bundle in bundle_index.values() if bundle.rcept_no})
    )
    results: List[Dict[str, Any]] = []
    issue_counts: Dict[str, int] = defaultdict(int)

    for item in sample:
        bundle = bundle_index.get(item["TXT_REDACTED"])
        if bundle is None:
            results.append({**item, "TXT_REDACTED": "TXT_REDACTED"})
            issue_counts["TXT_REDACTED"] += 4
            continue

        collector._current_store_asset_dirs = [bundle.asset_dir]
        current = collector._extract_fixed_asset_flows(bundle.rcept_no)
        candidates = _table_candidates(collector, bundle.rcept_no)
        candidates = _sort_candidates(candidates, item["TXT_REDACTED"], item["TXT_REDACTED"])
        issue = _issue_label(item, candidates)
        issue_counts[issue] += 1

        results.append(
            {
                **item,
                "TXT_REDACTED": bundle.company_key,
                "TXT_REDACTED": bundle.rcept_no,
                "TXT_REDACTED": current.get("TXT_REDACTED"),
                "TXT_REDACTED": current.get("TXT_REDACTED"),
                "TXT_REDACTED": _ratio(item["TXT_REDACTED"], item["TXT_REDACTED"]),
                "TXT_REDACTED": _ratio(item["TXT_REDACTED"], item["TXT_REDACTED"]),
                "TXT_REDACTED": issue,
                "TXT_REDACTED": len(candidates),
                "TXT_REDACTED": candidates[:2],
            }
        )

    payload = {
        "TXT_REDACTED": str(STORE_ROOT),
        "TXT_REDACTED": str(EXPECTED_XLSX),
        "TXT_REDACTED": str(PREDICTED_XLSX),
        "TXT_REDACTED": len(sample),
        "TXT_REDACTED": dict(sorted(issue_counts.items())),
        "TXT_REDACTED": results,
    }
    JSON_OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=3), encoding="TXT_REDACTED")

    lines = [
        "TXT_REDACTED",
        "TXT_REDACTED",
        "TXT_REDACTED"                               ,
        "TXT_REDACTED"                              ,
        "TXT_REDACTED"                                ,
        "TXT_REDACTED"                        ,
        "TXT_REDACTED",
        "TXT_REDACTED",
        "TXT_REDACTED",
    ]
    for issue, count in sorted(issue_counts.items()):
        lines.append("TXT_REDACTED"                       )
    lines.extend(["TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"])

    for result in results:
        lines.append("TXT_REDACTED"                                                 )
        lines.append("TXT_REDACTED"                                     )
        lines.append(
            "TXT_REDACTED"                                                                                 
            "TXT_REDACTED"                                                      
        )
        lines.append(
            "TXT_REDACTED"                                                                                    
            "TXT_REDACTED"                                                        
        )
        lines.append(
            "TXT_REDACTED"                                                                                                  
        )
        for candidate in result.get("TXT_REDACTED", [])[:4]:
            lines.append(
                "TXT_REDACTED"                                                              
                "TXT_REDACTED"                                                                         
                "TXT_REDACTED"                                                            
            )
        lines.append("TXT_REDACTED")

    MD_OUTPUT.write_text("TXT_REDACTED".join(lines), encoding="TXT_REDACTED")
    print("TXT_REDACTED"                    )
    print("TXT_REDACTED"                  )
    print(json.dumps(dict(sorted(issue_counts.items())), ensure_ascii=False))


if __name__ == "TXT_REDACTED":
    main()
