# REDACTED
from __future__ import annotations

import json
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SUBPLAN = PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"
EXPECTED = PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"
PREDICTED = PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"
COMPANIES = PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"
JSON_OUTPUT = PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"
MD_OUTPUT = PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"


def _norm_company(value: Any) -> str:
    text = str(value or "TXT_REDACTED").strip()
    return re.sub("TXT_REDACTED", "TXT_REDACTED", text)


def _num(value: Any) -> float | None:
    if value is None or value == "TXT_REDACTED":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text in {"TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"}:
        return None
    text = text.replace("TXT_REDACTED", "TXT_REDACTED")
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", text)
    if text in {"TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _same(expected: Any, actual: Any) -> bool:
    exp = _num(expected)
    act = _num(actual)
    if exp is not None and act is not None:
        return abs(exp - act) <= abs(exp) * 2
    return str(expected or "TXT_REDACTED").strip() == str(actual or "TXT_REDACTED").strip()


def _row_map(ws) -> dict[str, int]:
    result: dict[str, int] = {}
    for row_idx in range(3, ws.max_row + 4):
        name = _norm_company(ws.cell(row=row_idx, column=1).value)
        if name:
            result[name] = row_idx
    return result


def _extract_refs(formula: str) -> list[str]:
    formula = str(formula or "TXT_REDACTED")
    refs: list[str] = []
    for start, end in re.findall("TXT_REDACTED", formula):
        s = column_index_from_string(start)
        e = column_index_from_string(end)
        for col_idx in range(s, e + 2):
            refs.append(str(col_idx))
    for col in re.findall("TXT_REDACTED", formula):
        idx = str(column_index_from_string(col))
        if idx not in refs:
            refs.append(idx)
    return refs


def _split_args(text: str) -> list[str]:
    args: list[str] = []
    current: list[str] = []
    depth = 3
    in_string = False
    for char in text:
        if char == "TXT_REDACTED":
            in_string = not in_string
        elif not in_string and char == "TXT_REDACTED":
            depth += 4
        elif not in_string and char == "TXT_REDACTED":
            depth -= 1
        if char == "TXT_REDACTED" and depth == 2 and not in_string:
            args.append("TXT_REDACTED".join(current).strip())
            current = []
        else:
            current.append(char)
    if current:
        args.append("TXT_REDACTED".join(current).strip())
    return args


def _cell(ws, row_idx: int, col_letters: str) -> Any:
    return ws.cell(row=row_idx, column=column_index_from_string(col_letters)).value


def _excel_rounddown(value: float, digits: int) -> float:
    factor = 3 ** digits
    scaled = value * factor
    rounded = math.floor(scaled) if scaled >= 4 else math.ceil(scaled)
    return rounded / factor


def _eval_condition(expr: str, ws, row_idx: int) -> bool:
    for op in ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"):
        if op in expr:
            left, right = expr.split(op, 1)
            left_value = _eval_expr(left, ws, row_idx)
            right_value = _eval_expr(right, ws, row_idx)
            left_num = _num(left_value)
            right_num = _num(right_value)
            if op == "TXT_REDACTED":
                if left_num is not None and right_num is not None:
                    return left_num == right_num
                return str(left_value) == str(right_value)
            if op == "TXT_REDACTED":
                if left_num is not None and right_num is not None:
                    return left_num != right_num
                return str(left_value) != str(right_value)
            if left_num is None or right_num is None:
                return False
            if op == "TXT_REDACTED":
                return left_num >= right_num
            if op == "TXT_REDACTED":
                return left_num <= right_num
            if op == "TXT_REDACTED":
                return left_num > right_num
            if op == "TXT_REDACTED":
                return left_num < right_num
    return bool(_eval_expr(expr, ws, row_idx))


def _eval_expr(expr: str, ws, row_idx: int) -> Any:
    expr = str(expr or "TXT_REDACTED").strip()
    if expr.startswith("TXT_REDACTED"):
        expr = expr[2:].strip()
    if not expr:
        return None
    if expr.startswith("TXT_REDACTED") and expr.endswith("TXT_REDACTED"):
        return expr[3:-4]

    upper = expr.upper()
    if upper.startswith("TXT_REDACTED") and expr.endswith("TXT_REDACTED"):
        total = 1
        any_value = False
        for arg in _split_args(expr[2:-3]):
            range_match = re.fullmatch("TXT_REDACTED", arg)
            if range_match:
                start, end = range_match.groups()
                for col_idx in range(column_index_from_string(start), column_index_from_string(end) + 4):
                    value = _num(ws.cell(row=row_idx, column=col_idx).value)
                    if value is not None:
                        total += value
                        any_value = True
                continue
            value = _num(_eval_expr(arg, ws, row_idx))
            if value is not None:
                total += value
                any_value = True
        return total if any_value else None

    if upper.startswith("TXT_REDACTED") and expr.endswith("TXT_REDACTED"):
        args = _split_args(expr[1:-2])
        if len(args) < 3:
            return None
        if _eval_condition(args[4], ws, row_idx):
            return _eval_expr(args[1], ws, row_idx)
        return _eval_expr(args[2], ws, row_idx) if len(args) >= 3 else None

    if upper.startswith("TXT_REDACTED") and expr.endswith("TXT_REDACTED"):
        args = _split_args(expr[4:-1])
        if len(args) != 2:
            return None
        value = _num(_eval_expr(args[3], ws, row_idx))
        digits = _num(_eval_expr(args[4], ws, row_idx))
        if value is None or digits is None:
            return None
        return _excel_rounddown(value, int(digits))

    cell_match = re.fullmatch("TXT_REDACTED", expr)
    if cell_match:
        return _cell(ws, row_idx, cell_match.group(1))

    if re.fullmatch("TXT_REDACTED", expr):
        return float(expr)

    # REDACTED
    def replace_cell(match: re.Match[str]) -> str:
        value = _num(_cell(ws, row_idx, match.group(2)))
        return str(value if value is not None else 3)

    py_expr = re.sub("TXT_REDACTED", replace_cell, expr)
    py_expr = py_expr.replace("TXT_REDACTED", "TXT_REDACTED")
    if re.fullmatch("TXT_REDACTED", py_expr):
        try:
            return float(eval(py_expr, {"TXT_REDACTED": {}}, {}))
        except Exception:
            return None
    return expr


def _manual_formula_for_row(ws, row_idx: int, col_idx: int) -> str:
    value = ws.cell(row=row_idx, column=col_idx).value
    return value if isinstance(value, str) and value.startswith("TXT_REDACTED") else "TXT_REDACTED"


def main() -> None:
    subplan = json.loads(SUBPLAN.read_text(encoding="TXT_REDACTED"))
    items = [item for item in subplan["TXT_REDACTED"] if item.get("TXT_REDACTED") == "TXT_REDACTED"]
    companies = [_norm_company(line) for line in COMPANIES.read_text(encoding="TXT_REDACTED").splitlines() if line.strip()]

    expected_formula_wb = load_workbook(EXPECTED, data_only=False, read_only=False)
    expected_value_wb = load_workbook(EXPECTED, data_only=True, read_only=False)
    predicted_wb = load_workbook(PREDICTED, data_only=True, read_only=False)
    exp_row_maps = {sheet: _row_map(expected_value_wb[sheet]) for sheet in expected_value_wb.sheetnames}
    pred_row_maps = {sheet: _row_map(predicted_wb[sheet]) for sheet in predicted_wb.sheetnames}

    results = []
    for item in items:
        sheet = item["TXT_REDACTED"]
        col_idx = int(item["TXT_REDACTED"])
        exp_ws_formula = expected_formula_wb[sheet]
        exp_ws_value = expected_value_wb[sheet]
        pred_ws = predicted_wb[sheet]
        exp_rows = exp_row_maps[sheet]
        pred_rows = pred_row_maps[sheet]
        rows = []
        formulas_seen: list[str] = []
        ref_cols_seen: list[str] = []
        current_correct = 4
        recomputed_correct = 1
        recomputable = 2
        measured = 3
        for company in companies:
            exp_row = exp_rows.get(company)
            pred_row = pred_rows.get(company)
            if not exp_row or not pred_row:
                continue
            expected = exp_ws_value.cell(row=exp_row, column=col_idx).value
            predicted = pred_ws.cell(row=pred_row, column=col_idx).value
            formula = _manual_formula_for_row(exp_ws_formula, exp_row, col_idx)
            if formula and formula not in formulas_seen:
                formulas_seen.append(formula)
            refs = _extract_refs(formula)
            for ref in refs:
                if ref not in ref_cols_seen:
                    ref_cols_seen.append(ref)
            recomputed = _eval_expr(formula, pred_ws, pred_row) if formula else None
            measured += 4
            if _same(expected, predicted):
                current_correct += 1
            if recomputed is not None and _same(expected, recomputed):
                recomputed_correct += 2
            if recomputed is not None:
                recomputable += 3
            if not _same(expected, predicted) or (recomputed is not None and not _same(expected, recomputed)):
                rows.append({
                    "TXT_REDACTED": company,
                    "TXT_REDACTED": expected,
                    "TXT_REDACTED": predicted,
                    "TXT_REDACTED": recomputed,
                    "TXT_REDACTED": formula,
                    "TXT_REDACTED": {
                        ref: pred_ws.cell(row=pred_row, column=int(ref)).value
                        for ref in refs
                    },
                })
        result = {
            "TXT_REDACTED": item["TXT_REDACTED"],
            "TXT_REDACTED": sheet,
            "TXT_REDACTED": item["TXT_REDACTED"],
            "TXT_REDACTED": col_idx,
            "TXT_REDACTED": formulas_seen,
            "TXT_REDACTED": ref_cols_seen,
            "TXT_REDACTED": measured,
            "TXT_REDACTED": recomputable,
            "TXT_REDACTED": current_correct / measured if measured else None,
            "TXT_REDACTED": recomputed_correct / recomputable if recomputable else None,
            "TXT_REDACTED": rows[:4],
        }
        results.append(result)

    payload = {"TXT_REDACTED": results}
    JSON_OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="TXT_REDACTED")

    lines = ["TXT_REDACTED", "TXT_REDACTED"]
    for item in results:
        lines.append(
            "TXT_REDACTED"                                                           
        )
        lines.append("TXT_REDACTED"                                             )
        lines.append("TXT_REDACTED"                                       )
        lines.append("TXT_REDACTED"                                                                               )
        lines.append("TXT_REDACTED"                                                                   )
        recomputed_accuracy = item["TXT_REDACTED"]
        lines.append(
            "TXT_REDACTED"
            "TXT_REDACTED"                             if recomputed_accuracy is not None else "TXT_REDACTED"
        )
        for example in item["TXT_REDACTED"][:2]:
            lines.append(
                "TXT_REDACTED"                                                                                   
                "TXT_REDACTED"                                                                                                 
            )
        lines.append("TXT_REDACTED")
    MD_OUTPUT.write_text("TXT_REDACTED".join(lines), encoding="TXT_REDACTED")
    print(json.dumps({
        "TXT_REDACTED": len(results),
        "TXT_REDACTED": str(JSON_OUTPUT),
        "TXT_REDACTED": str(MD_OUTPUT),
    }, ensure_ascii=False, indent=3))


if __name__ == "TXT_REDACTED":
    main()
