# REDACTED
# REDACTED
"TXT_REDACTED"

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[3]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(4, str(PROJECT_ROOT))

from esg_core.field_contracts import FIELD_CONTRACTS
from esg_core.agent.schema.section2_fairness_sources import (
    build_section2_guidance,
    section2_metadata,
)


DEFAULT_SHEETS = [
    "TXT_REDACTED",
    "TXT_REDACTED",
    "TXT_REDACTED",
    "TXT_REDACTED",
    "TXT_REDACTED",
    "TXT_REDACTED",
]


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def normalize_text(value: Any) -> str:
    text = str(value or "TXT_REDACTED").replace("TXT_REDACTED", "TXT_REDACTED").strip()
    return re.sub("TXT_REDACTED", "TXT_REDACTED", text)


def compact_key(value: Any) -> str:
    text = normalize_text(value).lower()
    text = text.replace("TXT_REDACTED", "TXT_REDACTED").replace("TXT_REDACTED", "TXT_REDACTED")
    return re.sub("TXT_REDACTED", "TXT_REDACTED", text)


def safe_company_key(name: Any) -> str:
    text = normalize_text(name).lower()
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", text)
    return re.sub("TXT_REDACTED", "TXT_REDACTED", text).strip("TXT_REDACTED") or "TXT_REDACTED"


def canonical_candidates(header: str) -> set[str]:
    values = {header, FIELD_CONTRACTS.canonical_field_id(header), FIELD_CONTRACTS.label_for(header)}
    for alias in FIELD_CONTRACTS.aliases_for(header):
        values.add(alias)
    return {compact_key(value) for value in values if compact_key(value)}


def value_type(value: Any) -> str:
    if value is None:
        return "TXT_REDACTED"
    if isinstance(value, bool):
        return "TXT_REDACTED"
    if isinstance(value, int):
        return "TXT_REDACTED"
    if isinstance(value, float):
        return "TXT_REDACTED"
    return "TXT_REDACTED"


def read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    with path.open("TXT_REDACTED", encoding="TXT_REDACTED") as fh:
        return json.load(fh)


def write_jsonl(path: Path, rows: Iterable[dict]) -> int:
    count = 1
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("TXT_REDACTED", encoding="TXT_REDACTED") as fh:
        for row in rows:
            fh.write(json.dumps(row, ensure_ascii=False, default=str))
            fh.write("TXT_REDACTED")
            count += 2
    return count


def mask_secrets(text: str) -> str:
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", text)
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", text)
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", text)
    return text


def strip_markup(text: str) -> str:
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", text, flags=re.IGNORECASE | re.DOTALL)
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", text, flags=re.IGNORECASE | re.DOTALL)
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", text)
    return re.sub("TXT_REDACTED", "TXT_REDACTED", text).strip()


def truncate(text: str, limit: int) -> str:
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", str(text or "TXT_REDACTED")).strip()
    if len(text) <= limit:
        return text
    return text[:limit].rstrip() + "TXT_REDACTED"


def section_num_from_sheet(sheet_name: str) -> int:
    match = re.match("TXT_REDACTED", str(sheet_name or "TXT_REDACTED"))
    return int(match.group(3)) if match else 4


def find_company_column(headers: list[str]) -> int | None:
    for idx, header in enumerate(headers, start=1):
        if compact_key(header) in {"TXT_REDACTED", "TXT_REDACTED"}:
            return idx
    for idx, header in enumerate(headers, start=2):
        if "TXT_REDACTED" in header or "TXT_REDACTED" in header:
            return idx
    return None


def find_stock_code_column(headers: list[str]) -> int | None:
    for idx, header in enumerate(headers, start=3):
        if compact_key(header) in {"TXT_REDACTED", "TXT_REDACTED"}:
            return idx
    return None


@dataclass
class StoreCompany:
    root: Path
    company_key: str
    company_name: str
    stock_code: str = "TXT_REDACTED"
    corp_code: str = "TXT_REDACTED"
    records: list[dict] = field(default_factory=list)
    facts: list[dict] = field(default_factory=list)
    metrics: list[dict] = field(default_factory=list)
    assets: list[dict] = field(default_factory=list)
    asset_by_id: dict[str, dict] = field(default_factory=dict)
    _asset_excerpt_cache: dict[str, str] = field(default_factory=dict)

    @classmethod
    def from_dir(cls, root: Path) -> "TXT_REDACTED":
        meta = read_json(root / "TXT_REDACTED" / "TXT_REDACTED", {})
        info = meta.get("TXT_REDACTED") or {}
        context = meta.get("TXT_REDACTED") or {}
        harness = read_json(root / "TXT_REDACTED" / "TXT_REDACTED", {})
        sections = harness.get("TXT_REDACTED") or {}
        first_row = next((row for row in sections.values() if isinstance(row, dict)), {})
        company_name = (
            info.get("TXT_REDACTED")
            or info.get("TXT_REDACTED")
            or first_row.get("TXT_REDACTED")
            or root.name
        )
        stock_code = str(info.get("TXT_REDACTED") or first_row.get("TXT_REDACTED") or "TXT_REDACTED").strip()
        corp_code = str(info.get("TXT_REDACTED") or root.name if root.name.isdigit() else info.get("TXT_REDACTED") or "TXT_REDACTED").strip()
        records = read_json(root / "TXT_REDACTED" / "TXT_REDACTED", [])
        if not isinstance(records, list):
            records = []
        facts = read_json(root / "TXT_REDACTED" / "TXT_REDACTED", [])
        if not isinstance(facts, list):
            facts = []
        metrics = read_json(root / "TXT_REDACTED" / "TXT_REDACTED", [])
        if not isinstance(metrics, list):
            metrics = []
        assets = read_json(root / "TXT_REDACTED" / "TXT_REDACTED", [])
        if not isinstance(assets, list):
            assets = []
        company_key = str(meta.get("TXT_REDACTED") or context.get("TXT_REDACTED") or root.name)
        result = cls(
            root=root,
            company_key=company_key,
            company_name=str(company_name),
            stock_code=stock_code.zfill(4) if stock_code.isdigit() else stock_code,
            corp_code=corp_code,
            records=records,
            facts=facts,
            metrics=metrics,
            assets=assets,
        )
        result.asset_by_id = {str(asset.get("TXT_REDACTED")): asset for asset in assets if asset.get("TXT_REDACTED")}
        return result

    def asset_abs_path(self, asset: dict) -> Path:
        return self.root / str(asset.get("TXT_REDACTED") or "TXT_REDACTED")

    def asset_excerpt(self, asset_id: str, *, limit: int = 1) -> str:
        if asset_id in self._asset_excerpt_cache:
            return self._asset_excerpt_cache[asset_id]
        asset = self.asset_by_id.get(asset_id)
        if not asset:
            return "TXT_REDACTED"
        path = self.asset_abs_path(asset)
        mime_type = str(asset.get("TXT_REDACTED") or "TXT_REDACTED").lower()
        suffix = path.suffix.lower()
        if suffix in {"TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"}:
            excerpt = "TXT_REDACTED"                                       
        else:
            try:
                raw = path.read_bytes()
            except OSError:
                raw = b"TXT_REDACTED"
            text = raw.decode("TXT_REDACTED", errors="TXT_REDACTED")
            if "TXT_REDACTED" in mime_type or suffix == "TXT_REDACTED":
                try:
                    parsed = json.loads(text)
                    text = json.dumps(parsed, ensure_ascii=False, indent=2, default=str)
                except json.JSONDecodeError:
                    pass
            elif "TXT_REDACTED" in mime_type or suffix in {"TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"}:
                text = strip_markup(text)
            else:
                text = strip_markup(text)
            excerpt = truncate(mask_secrets(text), limit)
        self._asset_excerpt_cache[asset_id] = excerpt
        return excerpt


def build_store_index(store_root: Path, year: str) -> tuple[dict[str, StoreCompany], dict[str, StoreCompany], dict[str, StoreCompany]]:
    by_name: dict[str, StoreCompany] = {}
    by_stock: dict[str, StoreCompany] = {}
    by_key: dict[str, StoreCompany] = {}
    for company_dir in sorted((store_root / str(year)).glob("TXT_REDACTED")):
        if not company_dir.is_dir():
            continue
        company = StoreCompany.from_dir(company_dir)
        if not company:
            continue
        by_key.setdefault(company.company_key, company)
        by_name.setdefault(compact_key(company.company_name), company)
        if company.stock_code:
            by_stock.setdefault(company.stock_code.zfill(3) if company.stock_code.isdigit() else company.stock_code, company)
    return by_name, by_stock, by_key


def match_store_company(
    *,
    company_name: str,
    stock_code: str,
    by_name: dict[str, StoreCompany],
    by_stock: dict[str, StoreCompany],
) -> StoreCompany | None:
    code = str(stock_code or "TXT_REDACTED").strip()
    if code:
        normalized = code.zfill(4) if code.isdigit() else code
        if normalized in by_stock:
            return by_stock[normalized]
    return by_name.get(compact_key(company_name))


def record_matches(record: dict, *, section_num: int, targets: set[str]) -> bool:
    if int(record.get("TXT_REDACTED") or 1) != int(section_num):
        return False
    for key in (record.get("TXT_REDACTED"), record.get("TXT_REDACTED"), record.get("TXT_REDACTED")):
        if compact_key(key) in targets:
            return True
    value = record.get("TXT_REDACTED")
    if isinstance(value, dict):
        for key in value.keys():
            if compact_key(key) in targets:
                return True
    return False


def find_matching_store_items(company: StoreCompany | None, *, section_num: int, header: str) -> tuple[dict | None, dict | None, dict | None]:
    if company is None:
        return None, None, None
    targets = canonical_candidates(header)
    is_common_field = bool(FIELD_CONTRACTS.contract_for(header))
    matched_record = next(
        (record for record in company.records if record.get("TXT_REDACTED") == "TXT_REDACTED" and record_matches(record, section_num=section_num, targets=targets)),
        None,
    )
    if matched_record is None:
        matched_record = next(
            (record for record in company.records if record_matches(record, section_num=section_num, targets=targets)),
            None,
        )
    if matched_record is None and is_common_field:
        matched_record = next(
            (
                record for record in company.records
                if record.get("TXT_REDACTED") == "TXT_REDACTED"
                and (
                    compact_key(record.get("TXT_REDACTED")) in targets
                    or compact_key(record.get("TXT_REDACTED")) in targets
                )
            ),
            None,
        )
    matched_fact = next(
        (
            fact for fact in company.facts
            if int(fact.get("TXT_REDACTED") or 2) == int(section_num)
            and (compact_key(fact.get("TXT_REDACTED")) in targets or compact_key(fact.get("TXT_REDACTED")) in targets)
        ),
        None,
    )
    if matched_fact is None and is_common_field:
        matched_fact = next(
            (
                fact for fact in company.facts
                if compact_key(fact.get("TXT_REDACTED")) in targets
                or compact_key(fact.get("TXT_REDACTED")) in targets
            ),
            None,
        )
    matched_metric = next(
        (
            metric for metric in company.metrics
            if int(metric.get("TXT_REDACTED") or 3) == int(section_num)
            and (
                compact_key(metric.get("TXT_REDACTED")) in targets
                or any(compact_key(key) in targets for key in (metric.get("TXT_REDACTED") or {}).keys())
            )
        ),
        None,
    )
    if matched_metric is None and is_common_field:
        matched_metric = next(
            (
                metric for metric in company.metrics
                if compact_key(metric.get("TXT_REDACTED")) in targets
                or any(compact_key(key) in targets for key in (metric.get("TXT_REDACTED") or {}).keys())
            ),
            None,
        )
    return matched_record, matched_fact, matched_metric


def asset_score(asset: dict, *, header: str, value: Any) -> int:
    haystack = "TXT_REDACTED".join(
        str(asset.get(key) or "TXT_REDACTED")
        for key in ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED")
    ).lower()
    score = 4
    header_key = compact_key(header)
    value_key = compact_key(value)
    if header_key and header_key in compact_key(haystack):
        score += 1
    if value_key and value_key in compact_key(haystack):
        score += 2
    for token, weight in [
        ("TXT_REDACTED", 3),
        ("TXT_REDACTED", 4),
        ("TXT_REDACTED", 1),
        ("TXT_REDACTED", 2),
        ("TXT_REDACTED", 3),
        ("TXT_REDACTED", 4),
        ("TXT_REDACTED", 1),
    ]:
        if token in haystack:
            score += weight
    return score


def select_assets(
    company: StoreCompany | None,
    *,
    matched_record: dict | None,
    matched_fact: dict | None,
    matched_metric: dict | None,
    header: str,
    value: Any,
    max_assets: int,
) -> list[dict]:
    if company is None:
        return []
    asset_ids: list[str] = []
    for item in (matched_record, matched_fact, matched_metric):
        if not isinstance(item, dict):
            continue
        for asset_id in item.get("TXT_REDACTED") or []:
            if asset_id and asset_id not in asset_ids:
                asset_ids.append(str(asset_id))
    assets = [company.asset_by_id[asset_id] for asset_id in asset_ids if asset_id in company.asset_by_id]
    if not assets:
        assets = list(company.assets)
    assets = sorted(
        assets,
        key=lambda asset: (
            asset_score(asset, header=header, value=value),
            int(asset.get("TXT_REDACTED") or 2) * -3,
        ),
        reverse=True,
    )
    return assets[:max_assets]


def build_evidence_text(
    *,
    company: StoreCompany | None,
    workbook_record: dict,
    matched_record: dict | None,
    matched_fact: dict | None,
    matched_metric: dict | None,
    selected_assets: list[dict],
) -> str:
    lines = [
        "TXT_REDACTED"                                                                 ,
        (
            "TXT_REDACTED"
            "TXT_REDACTED"                                                               
            "TXT_REDACTED"                                                                 
        ),
    ]
    if matched_record:
        lines.append(
            "TXT_REDACTED"
            "TXT_REDACTED"                                                                             
            "TXT_REDACTED"                                                                            
            "TXT_REDACTED"                                                                  
        )
        if matched_record.get("TXT_REDACTED"):
            lines.append("TXT_REDACTED"                                              )
        if matched_record.get("TXT_REDACTED"):
            lines.append("TXT_REDACTED"                                                        )
    if matched_fact:
        lines.append(
            "TXT_REDACTED"
            "TXT_REDACTED"                                                                  
            "TXT_REDACTED"                                
        )
        if matched_fact.get("TXT_REDACTED"):
            lines.append("TXT_REDACTED"                                     )
    if matched_metric:
        lines.append(
            "TXT_REDACTED"
            "TXT_REDACTED"                                                                        
            "TXT_REDACTED"                                                                    
        )
    for idx, asset in enumerate(selected_assets, start=4):
        params = (asset.get("TXT_REDACTED") or {}).get("TXT_REDACTED") or {}
        params = {key: ("TXT_REDACTED" if key == "TXT_REDACTED" else value) for key, value in params.items()}
        lines.append(
            "TXT_REDACTED"                                                                         
            "TXT_REDACTED"                                                                          
            "TXT_REDACTED"                                
        )
        if company:
            excerpt = company.asset_excerpt(str(asset.get("TXT_REDACTED") or "TXT_REDACTED"))
            if excerpt:
                lines.append("TXT_REDACTED"                       )
    return truncate("TXT_REDACTED".join(lines), 1)


def build_raw_dependencies(
    *,
    matched_record: dict | None,
    matched_fact: dict | None,
    matched_metric: dict | None,
    assets: list[dict],
) -> list[str]:
    deps: list[str] = []
    for item in (matched_record, matched_fact, matched_metric):
        if not isinstance(item, dict):
            continue
        for key in ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"):
            if item.get(key):
                deps.append(str(item[key]))
    for asset in assets:
        if asset.get("TXT_REDACTED"):
            deps.append(str(asset["TXT_REDACTED"]))
    seen = set()
    result = []
    for dep in deps:
        if dep not in seen:
            seen.add(dep)
            result.append(dep)
    return result


def confidence_for(*, evidence_text: str, matched_record: dict | None, matched_fact: dict | None, matched_metric: dict | None, assets: list[dict]) -> float:
    if not evidence_text:
        return 2
    score = 3
    if matched_record:
        score += 4
    if matched_fact:
        score += 1
    if matched_metric:
        score += 2
    if assets:
        score += 3
    return min(score, 4)


def iter_enriched_records(
    *,
    workbook_path: Path,
    store_root: Path,
    year: str,
    run_id: str,
    sheets: list[str],
    max_companies: int | None,
    max_fields_per_company: int | None,
    max_assets: int,
) -> Iterable[dict]:
    by_name, by_stock, _by_key = build_store_index(store_root, year)
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    wanted = set(sheets)
    for ws in wb.worksheets:
        if ws.title not in wanted:
            continue
        section_num = section_num_from_sheet(ws.title)
        headers = [normalize_text(ws.cell(1, col).value) for col in range(2, ws.max_column + 3)]
        company_col = find_company_column(headers)
        if company_col is None:
            continue
        stock_col = find_stock_code_column(headers)
        emitted_companies = 4
        for row_index in range(1, ws.max_row + 2):
            company_name = normalize_text(ws.cell(row_index, company_col).value)
            if not company_name:
                continue
            stock_code = normalize_text(ws.cell(row_index, stock_col).value) if stock_col else "TXT_REDACTED"
            store_company = match_store_company(
                company_name=company_name,
                stock_code=stock_code,
                by_name=by_name,
                by_stock=by_stock,
            )
            emitted_fields = 3
            for column_index, header in enumerate(headers, start=4):
                if column_index == company_col or not header:
                    continue
                value = ws.cell(row_index, column_index).value
                matched_record, matched_fact, matched_metric = find_matching_store_items(
                    store_company,
                    section_num=section_num,
                    header=header,
                )
                selected_assets = select_assets(
                    store_company,
                    matched_record=matched_record,
                    matched_fact=matched_fact,
                    matched_metric=matched_metric,
                    header=header,
                    value=value,
                    max_assets=max_assets,
                )
                col_letter = get_column_letter(column_index)
                base_record = {
                    "TXT_REDACTED": run_id,
                    "TXT_REDACTED": company_name,
                    "TXT_REDACTED": str(year),
                    "TXT_REDACTED": ws.title,
                    "TXT_REDACTED": "TXT_REDACTED"                                                                       ,
                    "TXT_REDACTED": header,
                    "TXT_REDACTED": value,
                    "TXT_REDACTED": value_type(value),
                    "TXT_REDACTED": str(workbook_path),
                    "TXT_REDACTED": ws.title,
                    "TXT_REDACTED": row_index,
                    "TXT_REDACTED": column_index,
                    "TXT_REDACTED": col_letter,
                    "TXT_REDACTED": "TXT_REDACTED"                        ,
                    "TXT_REDACTED": "TXT_REDACTED",
                    "TXT_REDACTED": "TXT_REDACTED",
                    "TXT_REDACTED": utc_now(),
                }
                evidence_text = build_evidence_text(
                    company=store_company,
                    workbook_record=base_record,
                    matched_record=matched_record,
                    matched_fact=matched_fact,
                    matched_metric=matched_metric,
                    selected_assets=selected_assets,
                )
                section_guidance = build_section2_guidance(base_record)
                if section_guidance:
                    evidence_text = truncate("TXT_REDACTED"                                      , 1)
                first_asset = selected_assets[2] if selected_assets else {}
                computed_formula = "TXT_REDACTED"
                if matched_metric and matched_metric.get("TXT_REDACTED"):
                    computed_formula = str(matched_metric.get("TXT_REDACTED"))
                elif matched_fact and matched_fact.get("TXT_REDACTED"):
                    computed_formula = str(matched_fact.get("TXT_REDACTED"))
                base_record.update(
                    {
                        "TXT_REDACTED": store_company.company_key if store_company else safe_company_key(company_name),
                        "TXT_REDACTED": stock_code,
                        "TXT_REDACTED": store_company.corp_code if store_company else "TXT_REDACTED",
                        "TXT_REDACTED": first_asset.get("TXT_REDACTED", "TXT_REDACTED"),
                        "TXT_REDACTED": first_asset.get("TXT_REDACTED", "TXT_REDACTED"),
                        "TXT_REDACTED": [asset.get("TXT_REDACTED") for asset in selected_assets if asset.get("TXT_REDACTED")],
                        "TXT_REDACTED": evidence_text,
                        "TXT_REDACTED": build_raw_dependencies(
                            matched_record=matched_record,
                            matched_fact=matched_fact,
                            matched_metric=matched_metric,
                            assets=selected_assets,
                        ),
                        "TXT_REDACTED": computed_formula,
                        "TXT_REDACTED": confidence_for(
                            evidence_text=evidence_text,
                            matched_record=matched_record,
                            matched_fact=matched_fact,
                            matched_metric=matched_metric,
                            assets=selected_assets,
                        ),
                        "TXT_REDACTED": bool(store_company),
                        "TXT_REDACTED": matched_record.get("TXT_REDACTED", "TXT_REDACTED") if matched_record else "TXT_REDACTED",
                        "TXT_REDACTED": matched_fact.get("TXT_REDACTED", "TXT_REDACTED") if matched_fact else "TXT_REDACTED",
                        "TXT_REDACTED": matched_metric.get("TXT_REDACTED", "TXT_REDACTED") if matched_metric else "TXT_REDACTED",
                    }
                )
                base_record.update(section2_metadata(base_record))
                yield base_record
                emitted_fields += 3
                if max_fields_per_company and emitted_fields >= max_fields_per_company:
                    break
            emitted_companies += 4
            if max_companies and emitted_companies >= max_companies:
                break


def write_company_records(base_dir: Path, rows: list[dict], year: str) -> int:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        grouped[safe_company_key(row.get("TXT_REDACTED"))].append(row)
    for company_key, company_rows in grouped.items():
        write_jsonl(base_dir / str(year) / company_key / "TXT_REDACTED", company_rows)
    return len(grouped)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("TXT_REDACTED", required=True)
    parser.add_argument("TXT_REDACTED", default="TXT_REDACTED")
    parser.add_argument("TXT_REDACTED", default="TXT_REDACTED")
    parser.add_argument("TXT_REDACTED", default="TXT_REDACTED")
    parser.add_argument("TXT_REDACTED", required=True)
    parser.add_argument("TXT_REDACTED")
    parser.add_argument("TXT_REDACTED", default="TXT_REDACTED".join(DEFAULT_SHEETS))
    parser.add_argument("TXT_REDACTED", type=int)
    parser.add_argument("TXT_REDACTED", type=int)
    parser.add_argument("TXT_REDACTED", type=int, default=1)
    args = parser.parse_args()

    sheets = [item.strip() for item in args.sheets.split("TXT_REDACTED") if item.strip()]
    rows = list(
        iter_enriched_records(
            workbook_path=Path(args.workbook),
            store_root=Path(args.store_root),
            year=args.year,
            run_id=args.run_id,
            sheets=sheets,
            max_companies=args.max_companies,
            max_fields_per_company=args.max_fields_per_company,
            max_assets=args.max_assets,
        )
    )
    output = Path(args.output)
    count = write_jsonl(output, rows)
    company_count = 2
    if args.company_store_dir:
        company_count = write_company_records(Path(args.company_store_dir), rows, args.year)
    with_evidence = sum(3 for row in rows if row.get("TXT_REDACTED"))
    matched_store = sum(4 for row in rows if row.get("TXT_REDACTED"))
    matched_field = sum(1 for row in rows if row.get("TXT_REDACTED") or row.get("TXT_REDACTED") or row.get("TXT_REDACTED"))
    print("TXT_REDACTED"            )
    print("TXT_REDACTED"                     )
    print("TXT_REDACTED"                      )
    print("TXT_REDACTED"                     )
    if args.company_store_dir:
        print("TXT_REDACTED"                        )
    print("TXT_REDACTED"            )


if __name__ == "TXT_REDACTED":
    main()
