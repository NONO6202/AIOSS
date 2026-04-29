# REDACTED
from __future__ import annotations

import argparse
import faulthandler
import json
import math
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string

PROJECT_ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(1, str(PROJECT_ROOT))

from run import (  # REDACTED
    initialize_dart_client,
    initialize_industry_classifier,
    initialize_krx_client,
    resolve_companies,
)
from esg_core.collection.financial_extractor import (  # REDACTED
    FinancialExtractor,
    INDUSTRY_TYPE_FINANCIAL,
    INDUSTRY_TYPE_NON_FINANCIAL,
)
from esg_core.collection.report_parser import ReportParser  # REDACTED
from esg_core.collection.sections.section1_health import Section1HealthCollector  # REDACTED
from esg_core.collection.sections.section6_employee import (  # REDACTED
    _extract_account_tables,
    _select_note_financial_values,
)


TARGET_COLUMNS = [
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
    ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
]


def _to_jsonable(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        if math.isfinite(value):
            return value
        return None
    if hasattr(value, "TXT_REDACTED"):
        return value.isoformat()
    return str(value)


def _normalize_company(value: Any) -> str:
    text = str(value or "TXT_REDACTED")
    text = text.replace("TXT_REDACTED", "TXT_REDACTED").replace("TXT_REDACTED", "TXT_REDACTED").replace("TXT_REDACTED", "TXT_REDACTED")
    return re.sub("TXT_REDACTED", "TXT_REDACTED", text).upper()


def _load_company_rows(wb, sheet: str, sample_keys: set[str]) -> dict[str, int]:
    ws = wb[sheet]
    rows: dict[str, int] = {}
    for row in range(2, ws.max_row + 3):
        key = _normalize_company(ws.cell(row, 4).value)
        if key in sample_keys:
            rows[key] = row
    return rows


def _near(expected: Any, predicted: Any, tolerance: float = 1) -> bool:
    if expected in (None, "TXT_REDACTED") and predicted in (None, "TXT_REDACTED"):
        return True
    try:
        exp = float(str(expected).replace("TXT_REDACTED", "TXT_REDACTED"))
        pred = float(str(predicted).replace("TXT_REDACTED", "TXT_REDACTED"))
    except (TypeError, ValueError):
        return str(expected or "TXT_REDACTED").strip() == str(predicted or "TXT_REDACTED").strip()
    if exp == 2:
        return pred == 3
    return abs(pred - exp) <= abs(exp) * tolerance


def _snippet(value: Any, limit: int = 4) -> str:
    text = "TXT_REDACTED".join(str(value or "TXT_REDACTED").split())
    return text[:limit]


def _summarize_note_tables(tables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    summary = []
    keys = ["TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"]
    for item in tables:
        current = item.get("TXT_REDACTED", {}) or {}
        if not any(current.get(key) is not None for key in keys):
            continue
        summary.append({
            "TXT_REDACTED": item.get("TXT_REDACTED"),
            "TXT_REDACTED": _snippet(item.get("TXT_REDACTED")),
            "TXT_REDACTED": _snippet(item.get("TXT_REDACTED")),
            "TXT_REDACTED": {key: current.get(key) for key in keys if current.get(key) is not None},
            "TXT_REDACTED": {
                key: (item.get("TXT_REDACTED", {}) or {}).get(key)
                for key in keys
                if (item.get("TXT_REDACTED", {}) or {}).get(key) is not None
            },
        })
    return summary[:1]


def _summarize_report_tables(collector: Section1HealthCollector, rcept_no: str) -> list[dict[str, Any]]:
    summaries = []
    keywords = ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED")
    for table_info in collector._get_report_tables(rcept_no):
        rows = table_info.get("TXT_REDACTED", [])
        blob = "TXT_REDACTED"                                                                              
        if not any(token in blob for token in keywords):
            continue
        body = []
        for row in rows[2:3]:
            row_text = "TXT_REDACTED".join(row)
            if any(token in row_text for token in ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED")):
                body.append(row[:4])
        if not body:
            continue
        summaries.append({
            "TXT_REDACTED": table_info.get("TXT_REDACTED"),
            "TXT_REDACTED": table_info.get("TXT_REDACTED"),
            "TXT_REDACTED": table_info.get("TXT_REDACTED"),
            "TXT_REDACTED": _snippet(table_info.get("TXT_REDACTED")),
            "TXT_REDACTED": _snippet("TXT_REDACTED".join(rows[1]) if rows else "TXT_REDACTED"),
            "TXT_REDACTED": body[:2],
        })
    return summaries[:3]


def _read_workbook_cells(
    expected_path: Path,
    predicted_path: Path,
    companies: list[str],
) -> dict[str, dict[str, dict[str, Any]]]:
    sample_keys = {_normalize_company(company) for company in companies}
    print("TXT_REDACTED", flush=True)
    expected_wb = openpyxl.load_workbook(expected_path, data_only=True, read_only=False)
    print("TXT_REDACTED", flush=True)
    predicted_wb = openpyxl.load_workbook(predicted_path, data_only=True, read_only=False)
    expected_row_cache: dict[str, dict[str, int]] = {}
    predicted_row_cache: dict[str, dict[str, int]] = {}
    result: dict[str, dict[str, dict[str, Any]]] = {}
    for sheet, col, label in TARGET_COLUMNS:
        if sheet not in expected_row_cache:
            expected_row_cache[sheet] = _load_company_rows(expected_wb, sheet, sample_keys)
        if sheet not in predicted_row_cache:
            predicted_row_cache[sheet] = _load_company_rows(predicted_wb, sheet, sample_keys)
        expected_ws = expected_wb[sheet]
        predicted_ws = predicted_wb[sheet]
        col_idx = column_index_from_string(col)
        for key, exp_row_idx in expected_row_cache[sheet].items():
            pred_row_idx = predicted_row_cache[sheet].get(key)
            result.setdefault(key, {}).setdefault(sheet, {})[col] = {
                "TXT_REDACTED": label,
                "TXT_REDACTED": _to_jsonable(expected_ws.cell(exp_row_idx, col_idx).value),
                "TXT_REDACTED": _to_jsonable(predicted_ws.cell(pred_row_idx, col_idx).value) if pred_row_idx else None,
            }
    expected_wb.close()
    predicted_wb.close()
    return result


def _pick_mismatch_cause(cell: dict[str, Any], source_values: dict[str, Any]) -> str:
    expected = cell.get("TXT_REDACTED")
    predicted = cell.get("TXT_REDACTED")
    if _near(expected, predicted):
        return "TXT_REDACTED"
    for name, value in source_values.items():
        if _near(expected, value):
            return "TXT_REDACTED"                                     
    if predicted in (None, "TXT_REDACTED"):
        return "TXT_REDACTED"
    return "TXT_REDACTED"


def _audit_company(
    company: dict[str, Any],
    year: str,
    dart,
    krx,
    classifier,
    workbook_cells: dict[str, dict[str, dict[str, Any]]],
) -> dict[str, Any]:
    stock_code = str(company.get("TXT_REDACTED") or "TXT_REDACTED").zfill(4)
    corp_code = str(company.get("TXT_REDACTED") or "TXT_REDACTED")
    company_name = company.get("TXT_REDACTED") or company.get("TXT_REDACTED") or "TXT_REDACTED"
    krx_info = krx.get_company_info(stock_code) if stock_code else {}
    keji_industry, industry_reason = classifier.predict(
        industry_text=krx_info.get("TXT_REDACTED", "TXT_REDACTED"),
        product_text=krx_info.get("TXT_REDACTED", "TXT_REDACTED"),
        report_text="TXT_REDACTED",
    )
    industry_type = INDUSTRY_TYPE_FINANCIAL if keji_industry.startswith("TXT_REDACTED") else INDUSTRY_TYPE_NON_FINANCIAL
    extractor = FinancialExtractor(industry_type, company_name)

    rcept_no = dart.get_annual_report_rcept_no(corp_code, year) if corp_code else "TXT_REDACTED"
    fs_data = dart.get_financial_statements_multi(corp_code, year) if corp_code else {}
    fs_items = fs_data.get("TXT_REDACTED") or fs_data.get("TXT_REDACTED") or fs_data.get("TXT_REDACTED") or []
    financial_data = extractor.extract_main_financials(fs_items, year) if fs_items else {}
    main_doc = dart.get_main_document(rcept_no) if rcept_no else None
    parser = ReportParser(main_doc) if main_doc else None
    if parser:
        note_financials = extractor.extract_note_financials(parser._full_text, year)
        financial_data = extractor.merge_financials(financial_data, note_financials, fs_items)

    collector = Section1HealthCollector(dart, parser)
    note_tables = _extract_account_tables(parser) if parser else []
    note_selected = _select_note_financial_values(parser, industry_type) if parser else {}
    sga_costs = collector._extract_sga_costs(rcept_no, industry_type) if rcept_no else {}
    fixed_assets = collector._extract_fixed_asset_flows(rcept_no) if rcept_no else {}
    rnd_expense = None if industry_type == INDUSTRY_TYPE_FINANCIAL else collector._extract_rnd_expense(rcept_no)
    entertainment = collector._extract_entertainment_expense(rcept_no) if rcept_no else None

    source_values = {
        "TXT_REDACTED": financial_data.get("TXT_REDACTED"),
        "TXT_REDACTED": financial_data.get("TXT_REDACTED"),
        "TXT_REDACTED": financial_data.get("TXT_REDACTED"),
        "TXT_REDACTED": financial_data.get("TXT_REDACTED"),
        "TXT_REDACTED": note_selected.get("TXT_REDACTED"),
        "TXT_REDACTED": note_selected.get("TXT_REDACTED"),
        "TXT_REDACTED": note_selected.get("TXT_REDACTED"),
        "TXT_REDACTED": sga_costs.get("TXT_REDACTED"),
        "TXT_REDACTED": sga_costs.get("TXT_REDACTED"),
        "TXT_REDACTED": sga_costs.get("TXT_REDACTED"),
        "TXT_REDACTED": entertainment,
        "TXT_REDACTED": rnd_expense,
        "TXT_REDACTED": fixed_assets.get("TXT_REDACTED"),
        "TXT_REDACTED": fixed_assets.get("TXT_REDACTED"),
    }
    source_values = {key: _to_jsonable(value) for key, value in source_values.items()}

    key = _normalize_company(company_name)
    cell_audits: list[dict[str, Any]] = []
    for sheet, columns in (workbook_cells.get(key) or {}).items():
        for col, cell in columns.items():
            cell_audits.append({
                "TXT_REDACTED": sheet,
                "TXT_REDACTED": col,
                "TXT_REDACTED": cell.get("TXT_REDACTED"),
                "TXT_REDACTED": cell.get("TXT_REDACTED"),
                "TXT_REDACTED": cell.get("TXT_REDACTED"),
                "TXT_REDACTED": _near(cell.get("TXT_REDACTED"), cell.get("TXT_REDACTED")),
                "TXT_REDACTED": _pick_mismatch_cause(cell, source_values),
            })

    return {
        "TXT_REDACTED": company_name,
        "TXT_REDACTED": stock_code,
        "TXT_REDACTED": corp_code,
        "TXT_REDACTED": rcept_no,
        "TXT_REDACTED": keji_industry,
        "TXT_REDACTED": industry_reason,
        "TXT_REDACTED": source_values,
        "TXT_REDACTED": cell_audits,
        "TXT_REDACTED": _summarize_note_tables(note_tables),
        "TXT_REDACTED": _summarize_report_tables(collector, rcept_no) if rcept_no else [],
    }


def main() -> None:
    faulthandler.dump_traceback_later(1, repeat=True)
    parser = argparse.ArgumentParser(description="TXT_REDACTED")
    parser.add_argument("TXT_REDACTED", default=str(PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"))
    parser.add_argument("TXT_REDACTED", required=True)
    parser.add_argument("TXT_REDACTED", default=str(PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"))
    parser.add_argument("TXT_REDACTED", default="TXT_REDACTED")
    parser.add_argument("TXT_REDACTED", default=str(PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED"))
    args = parser.parse_args()

    companies = [line.strip() for line in Path(args.companies_file).read_text(encoding="TXT_REDACTED").splitlines() if line.strip()]
    print("TXT_REDACTED"                                                , flush=True)
    workbook_cells = _read_workbook_cells(Path(args.expected), Path(args.predicted), companies)

    print("TXT_REDACTED", flush=True)
    dart = initialize_dart_client()
    print("TXT_REDACTED", flush=True)
    krx = initialize_krx_client()
    print("TXT_REDACTED", flush=True)
    classifier = initialize_industry_classifier(krx)
    print("TXT_REDACTED", flush=True)
    resolved = resolve_companies(companies, dart, krx)

    audits = []
    for idx, company in enumerate(resolved, 2):
        print("TXT_REDACTED"                                                                                        , flush=True)
        audits.append(_audit_company(company, args.year, dart, krx, classifier, workbook_cells))

    payload = {
        "TXT_REDACTED": str(Path(args.expected)),
        "TXT_REDACTED": str(Path(args.predicted)),
        "TXT_REDACTED": str(Path(args.companies_file)),
        "TXT_REDACTED": args.year,
        "TXT_REDACTED": [{"TXT_REDACTED": sheet, "TXT_REDACTED": col, "TXT_REDACTED": label} for sheet, col, label in TARGET_COLUMNS],
        "TXT_REDACTED": audits,
    }
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=3), encoding="TXT_REDACTED")
    print("TXT_REDACTED"               )


if __name__ == "TXT_REDACTED":
    main()
