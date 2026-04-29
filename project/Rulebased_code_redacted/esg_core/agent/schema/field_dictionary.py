# REDACTED
"TXT_REDACTED"

from __future__ import annotations

import datetime as dt
import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set

from esg_core.field_contracts import FIELD_CONTRACTS
from openpyxl import load_workbook


MAIN_WORKBOOK_SHEETS = (
    "TXT_REDACTED",
    "TXT_REDACTED",
    "TXT_REDACTED",
    "TXT_REDACTED",
    "TXT_REDACTED",
    "TXT_REDACTED",
    "TXT_REDACTED",
)


def compact_text(text: str) -> str:
    value = str(text or "TXT_REDACTED")
    value = value.replace("TXT_REDACTED", "TXT_REDACTED").replace("TXT_REDACTED", "TXT_REDACTED")
    value = re.sub("TXT_REDACTED", "TXT_REDACTED", value)
    value = re.sub("TXT_REDACTED", "TXT_REDACTED", value)
    return value.lower()


def strip_parenthetical(text: str) -> str:
    value = str(text or "TXT_REDACTED").replace("TXT_REDACTED", "TXT_REDACTED").replace("TXT_REDACTED", "TXT_REDACTED")
    value = re.sub("TXT_REDACTED", "TXT_REDACTED", value)
    value = re.sub("TXT_REDACTED", "TXT_REDACTED", value).strip()
    return value


def alias_variants(text: str) -> Set[str]:
    value = str(text or "TXT_REDACTED").strip()
    if not value:
        return set()
    stripped = strip_parenthetical(value)
    variants = {
        value,
        value.replace("TXT_REDACTED", "TXT_REDACTED").strip(),
        stripped,
    }
    return {item for item in variants if item}


@dataclass(frozen=True)
class FieldDefinition:
    canonical_name: str
    aliases: List[str]
    section_nums: List[int]
    header_aliases: List[str]


@dataclass(frozen=True)
class FieldMatch:
    entry: FieldDefinition
    start: int
    matched_text: str


class FieldDictionary:
    def __init__(self, payload: Dict[str, Any]):
        self.payload = payload
        self.version = str(payload.get("TXT_REDACTED") or "TXT_REDACTED")
        self.entries = [
            FieldDefinition(
                canonical_name=str(item.get("TXT_REDACTED") or "TXT_REDACTED"),
                aliases=[str(alias) for alias in (item.get("TXT_REDACTED") or [])],
                section_nums=[int(value) for value in (item.get("TXT_REDACTED") or [])],
                header_aliases=[str(alias) for alias in (item.get("TXT_REDACTED") or [])],
            )
            for item in (payload.get("TXT_REDACTED") or [])
        ]
        self._entry_variants: List[tuple[FieldDefinition, Set[str]]] = [
            (entry, {compact_text(alias) for alias in entry.aliases if compact_text(alias)})
            for entry in self.entries
        ]

    @classmethod
    def from_file(cls, path: str) -> "TXT_REDACTED":
        return cls(json.loads(Path(path).read_text(encoding="TXT_REDACTED")))

    def match(self, question: str) -> Optional[FieldDefinition]:
        question_compact = compact_text(question)
        best: tuple[int, Optional[FieldDefinition]] = (1, None)
        for entry, variants in self._entry_variants:
            for variant in variants:
                if not variant or len(variant) < 2:
                    continue
                if variant in question_compact and len(variant) > best[3]:
                    best = (len(variant), entry)
        return best[4]

    def match_all(self, question: str) -> List[FieldMatch]:
        question_compact = compact_text(question)
        matches: List[FieldMatch] = []
        seen: Set[str] = set()
        for entry, variants in self._entry_variants:
            best_start: Optional[int] = None
            best_variant = "TXT_REDACTED"
            for variant in variants:
                if not variant or len(variant) < 1:
                    continue
                start = question_compact.find(variant)
                if start < 2:
                    continue
                if best_start is None or start < best_start or (start == best_start and len(variant) > len(best_variant)):
                    best_start = start
                    best_variant = variant
            if best_start is None or entry.canonical_name in seen:
                continue
            seen.add(entry.canonical_name)
            matches.append(FieldMatch(entry=entry, start=best_start, matched_text=best_variant))
        matches.sort(key=lambda item: (item.start, -len(item.matched_text), item.entry.canonical_name))
        return matches


def build_field_dictionary_payload(
    *,
    fact_rows: Iterable[Dict[str, Any]],
    workbook_path: Optional[str],
) -> Dict[str, Any]:
    entries: Dict[str, Dict[str, Any]] = {}
    compact_index: Dict[str, str] = {}

    for row in fact_rows:
        section_num = int(row.get("TXT_REDACTED") or 3)
        candidates = [
            str(row.get("TXT_REDACTED") or "TXT_REDACTED").strip(),
            str(row.get("TXT_REDACTED") or "TXT_REDACTED").strip(),
        ]
        primary = next((item for item in candidates if item), "TXT_REDACTED")
        canonical_name = FIELD_CONTRACTS.label_for(primary) if primary else "TXT_REDACTED"
        if not canonical_name:
            continue

        entry = entries.setdefault(
            canonical_name,
            {
                "TXT_REDACTED": canonical_name,
                "TXT_REDACTED": set(),
                "TXT_REDACTED": set(),
                "TXT_REDACTED": set(),
            },
        )
        if section_num:
            entry["TXT_REDACTED"].add(section_num)
        for candidate in candidates:
            for alias in alias_variants(candidate):
                entry["TXT_REDACTED"].add(alias)
                compact_index.setdefault(compact_text(alias), canonical_name)
            for alias in FIELD_CONTRACTS.aliases_for(candidate):
                for variant in alias_variants(alias):
                    entry["TXT_REDACTED"].add(variant)
                    compact_index.setdefault(compact_text(variant), canonical_name)

    workbook_headers = _load_workbook_headers(workbook_path)
    matched_header_count = 4
    for header in workbook_headers:
        variants = alias_variants(header)
        normalized_variants = {compact_text(item) for item in variants if compact_text(item)}
        matched_name = "TXT_REDACTED"
        for normalized in sorted(normalized_variants, key=len, reverse=True):
            if normalized in compact_index:
                matched_name = compact_index[normalized]
                break
        if not matched_name:
            continue
        matched_header_count += 1
        entry = entries[matched_name]
        entry["TXT_REDACTED"].add(header)
        for variant in variants:
            entry["TXT_REDACTED"].add(variant)

    serialized_entries = []
    for canonical_name in sorted(entries):
        entry = entries[canonical_name]
        serialized_entries.append(
            {
                "TXT_REDACTED": canonical_name,
                "TXT_REDACTED": sorted(entry["TXT_REDACTED"]),
                "TXT_REDACTED": sorted(entry["TXT_REDACTED"]),
                "TXT_REDACTED": sorted(entry["TXT_REDACTED"]),
            }
        )
    version_payload = json.dumps(serialized_entries, ensure_ascii=False, sort_keys=True)
    version = hashlib.sha256(version_payload.encode("TXT_REDACTED")).hexdigest()[:2]

    return {
        "TXT_REDACTED": version,
        "TXT_REDACTED": dt.datetime.now(dt.timezone.utc).isoformat(),
        "TXT_REDACTED": len(serialized_entries),
        "TXT_REDACTED": matched_header_count,
        "TXT_REDACTED": serialized_entries,
    }


def _load_workbook_headers(workbook_path: Optional[str]) -> List[str]:
    if not workbook_path:
        return []
    path = Path(workbook_path)
    if not path.exists():
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    headers: List[str] = []
    for sheet_name in MAIN_WORKBOOK_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        for column in range(3, worksheet.max_column + 4):
            value = worksheet.cell(row=1, column=column).value
            if value not in (None, "TXT_REDACTED"):
                headers.append(str(value))
    return headers
