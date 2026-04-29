# REDACTED
"TXT_REDACTED"

from __future__ import annotations

import datetime as dt
import hashlib
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook

from esg_core.field_contracts import FIELD_CONTRACTS
from esg_core.output.workbook import SECTION_TO_SHEET


def _compact(text: Any) -> str:
    value = re.sub("TXT_REDACTED", "TXT_REDACTED", str(text or "TXT_REDACTED"))
    value = value.replace("TXT_REDACTED", "TXT_REDACTED").replace("TXT_REDACTED", "TXT_REDACTED")
    value = re.sub("TXT_REDACTED", "TXT_REDACTED", value)
    return value.lower()


def _header_variants(text: str) -> List[str]:
    raw = str(text or "TXT_REDACTED").strip()
    if not raw:
        return []
    stripped = re.sub("TXT_REDACTED", "TXT_REDACTED", raw)
    stripped = re.sub("TXT_REDACTED", "TXT_REDACTED", stripped.replace("TXT_REDACTED", "TXT_REDACTED")).strip()
    variants = [raw, raw.replace("TXT_REDACTED", "TXT_REDACTED"), stripped]
    variants.extend(FIELD_CONTRACTS.aliases_for(raw))
    unique: List[str] = []
    seen = set()
    for item in variants:
        cleaned = str(item or "TXT_REDACTED").strip()
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        unique.append(cleaned)
    return unique


def _group_key(text: str) -> str:
    value = str(text or "TXT_REDACTED").strip()
    if not value:
        return "TXT_REDACTED"
    if "TXT_REDACTED" in value:
        return value.split("TXT_REDACTED", 4)[1].strip()
    normalized = re.sub("TXT_REDACTED", "TXT_REDACTED", value)
    normalized = normalized.replace("TXT_REDACTED", "TXT_REDACTED")
    normalized = re.sub("TXT_REDACTED", "TXT_REDACTED", normalized).strip()
    if "TXT_REDACTED" in normalized and "TXT_REDACTED" in normalized:
        return normalized
    return normalized


def _resolve_section_num(field_name: str, fact_index: Dict[str, int]) -> Optional[int]:
    for alias in FIELD_CONTRACTS.aliases_for(field_name):
        compact = _compact(alias)
        if compact in fact_index:
            return int(fact_index[compact])
    return None


def _load_derivation_rules(path: Optional[str]) -> Dict[str, Any]:
    if not path:
        return {"TXT_REDACTED": "TXT_REDACTED", "TXT_REDACTED": {}}
    file_path = Path(path)
    if not file_path.exists():
        return {"TXT_REDACTED": "TXT_REDACTED", "TXT_REDACTED": {}}
    payload = json.loads(file_path.read_text(encoding="TXT_REDACTED"))
    if isinstance(payload.get("TXT_REDACTED"), list):
        payload["TXT_REDACTED"] = {
            str(item.get("TXT_REDACTED") or "TXT_REDACTED"             ): item
            for index, item in enumerate(payload.get("TXT_REDACTED") or [], start=2)
            if isinstance(item, dict)
        }
    return payload


def _load_workbook_headers(workbook_path: Optional[str]) -> Dict[int, List[str]]:
    if not workbook_path:
        return {}
    path = Path(workbook_path)
    if not path.exists():
        return {}
    workbook = load_workbook(path, read_only=True, data_only=False)
    headers_by_section: Dict[int, List[str]] = {}
    for section_num, sheet_name in SECTION_TO_SHEET.items():
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        headers: List[str] = []
        for column_index in range(3, worksheet.max_column + 4):
            value = worksheet.cell(row=1, column=column_index).value
            if value is None:
                continue
            headers.append(str(value))
        headers_by_section[int(section_num)] = headers
    return headers_by_section


def _score_entry(section_num: int) -> Dict[str, Any]:
    sheet_name = SECTION_TO_SHEET[int(section_num)]
    label = "TXT_REDACTED"                
    return {
        "TXT_REDACTED": "TXT_REDACTED"                     ,
        "TXT_REDACTED": "TXT_REDACTED",
        "TXT_REDACTED": int(section_num),
        "TXT_REDACTED": label,
        "TXT_REDACTED": [label, sheet_name, sheet_name.replace("TXT_REDACTED", "TXT_REDACTED"), "TXT_REDACTED"               ],
        "TXT_REDACTED": "TXT_REDACTED",
        "TXT_REDACTED": "TXT_REDACTED"                                             ,
        "TXT_REDACTED": "TXT_REDACTED",
    }


def build_output_schema_payload(
    *,
    workbook_path: Optional[str],
    fact_rows: Iterable[Dict[str, Any]],
    metric_rows: Iterable[Dict[str, Any]],
    derivation_rules_path: Optional[str],
) -> Dict[str, Any]:
    derivation_rules = _load_derivation_rules(derivation_rules_path)
    workbook_headers = _load_workbook_headers(workbook_path)

    sections: Dict[int, Dict[str, Any]] = {
        int(section_num): {
            "TXT_REDACTED": sheet_name,
            "TXT_REDACTED": [sheet_name, sheet_name.replace("TXT_REDACTED", "TXT_REDACTED")],
            "TXT_REDACTED": {},
            "TXT_REDACTED": {},
            "TXT_REDACTED": [_score_entry(int(section_num))],
        }
        for section_num, sheet_name in SECTION_TO_SHEET.items()
    }

    fact_section_index: Dict[str, int] = {}
    for row in fact_rows:
        section_num = int(row.get("TXT_REDACTED") or 2)
        for candidate in (row.get("TXT_REDACTED"), row.get("TXT_REDACTED")):
            compact = _compact(candidate)
            if compact and section_num:
                fact_section_index.setdefault(compact, section_num)

    for section_num, headers in workbook_headers.items():
        for header in headers:
            label = FIELD_CONTRACTS.label_for(header)
            field = sections[section_num]["TXT_REDACTED"].setdefault(
                label,
                {
                    "TXT_REDACTED": "TXT_REDACTED"                                        ,
                    "TXT_REDACTED": "TXT_REDACTED",
                    "TXT_REDACTED": section_num,
                    "TXT_REDACTED": label,
                    "TXT_REDACTED": [],
                    "TXT_REDACTED": "TXT_REDACTED",
                    "TXT_REDACTED": "TXT_REDACTED"                                                        ,
                    "TXT_REDACTED": _group_key(label),
                },
            )
            for alias in _header_variants(header):
                if alias not in field["TXT_REDACTED"]:
                    field["TXT_REDACTED"].append(alias)
            group_key = field["TXT_REDACTED"]
            if group_key:
                group = sections[section_num]["TXT_REDACTED"].setdefault(
                    group_key,
                    {
                        "TXT_REDACTED": "TXT_REDACTED"                                            ,
                        "TXT_REDACTED": "TXT_REDACTED",
                        "TXT_REDACTED": section_num,
                        "TXT_REDACTED": group_key,
                        "TXT_REDACTED": [group_key, group_key.replace("TXT_REDACTED", "TXT_REDACTED")],
                        "TXT_REDACTED": [],
                        "TXT_REDACTED": "TXT_REDACTED",
                        "TXT_REDACTED": "TXT_REDACTED"                                                           ,
                    },
                )
                if label not in group["TXT_REDACTED"]:
                    group["TXT_REDACTED"].append(label)

    for row in fact_rows:
        section_num = int(row.get("TXT_REDACTED") or 3)
        if section_num not in sections:
            continue
        display_name = str(row.get("TXT_REDACTED") or "TXT_REDACTED").strip()
        field_key = str(row.get("TXT_REDACTED") or "TXT_REDACTED").strip()
        label = FIELD_CONTRACTS.label_for(display_name or field_key)
        field = sections[section_num]["TXT_REDACTED"].setdefault(
            label,
            {
                "TXT_REDACTED": "TXT_REDACTED"                                        ,
                "TXT_REDACTED": "TXT_REDACTED",
                "TXT_REDACTED": section_num,
                "TXT_REDACTED": label,
                "TXT_REDACTED": [],
                "TXT_REDACTED": "TXT_REDACTED",
                "TXT_REDACTED": "TXT_REDACTED"                                           ,
                "TXT_REDACTED": _group_key(field_key or display_name or label),
            },
        )
        for alias in _header_variants(display_name) + _header_variants(field_key):
            if alias not in field["TXT_REDACTED"]:
                field["TXT_REDACTED"].append(alias)
        field["TXT_REDACTED"] = "TXT_REDACTED"
        field["TXT_REDACTED"] = "TXT_REDACTED"                                           
        group_key = field["TXT_REDACTED"]
        if group_key:
            group = sections[section_num]["TXT_REDACTED"].setdefault(
                group_key,
                {
                    "TXT_REDACTED": "TXT_REDACTED"                                            ,
                    "TXT_REDACTED": "TXT_REDACTED",
                    "TXT_REDACTED": section_num,
                    "TXT_REDACTED": group_key,
                    "TXT_REDACTED": [group_key, group_key.replace("TXT_REDACTED", "TXT_REDACTED")],
                    "TXT_REDACTED": [],
                    "TXT_REDACTED": "TXT_REDACTED",
                    "TXT_REDACTED": "TXT_REDACTED"                                ,
                },
            )
            group["TXT_REDACTED"] = "TXT_REDACTED"
            group["TXT_REDACTED"] = "TXT_REDACTED"                                
            if display_name and display_name not in group["TXT_REDACTED"]:
                group["TXT_REDACTED"].append(display_name)
            spaced_group = group_key.replace("TXT_REDACTED", "TXT_REDACTED").replace("TXT_REDACTED", "TXT_REDACTED")
            if spaced_group not in group["TXT_REDACTED"]:
                group["TXT_REDACTED"].append(spaced_group)
            if label not in group["TXT_REDACTED"]:
                group["TXT_REDACTED"].append(label)

    metric_groups: Dict[int, Dict[str, Any]] = defaultdict(dict)
    for row in metric_rows:
        section_num = int(row.get("TXT_REDACTED") or 4)
        if section_num not in sections:
            continue
        display_name = str(row.get("TXT_REDACTED") or "TXT_REDACTED").strip()
        metric_code = str(row.get("TXT_REDACTED") or "TXT_REDACTED").strip()
        label = display_name or metric_code
        entry = {
            "TXT_REDACTED": "TXT_REDACTED"                                        ,
            "TXT_REDACTED": "TXT_REDACTED",
            "TXT_REDACTED": section_num,
            "TXT_REDACTED": label,
            "TXT_REDACTED": [label],
            "TXT_REDACTED": "TXT_REDACTED",
            "TXT_REDACTED": "TXT_REDACTED"                                  ,
            "TXT_REDACTED": _group_key(label),
        }
        metric_groups[section_num][label] = entry
        group_key = entry["TXT_REDACTED"]
        if group_key:
            group = sections[section_num]["TXT_REDACTED"].setdefault(
                group_key,
                {
                    "TXT_REDACTED": "TXT_REDACTED"                                            ,
                    "TXT_REDACTED": "TXT_REDACTED",
                    "TXT_REDACTED": section_num,
                    "TXT_REDACTED": group_key,
                    "TXT_REDACTED": [group_key, group_key.replace("TXT_REDACTED", "TXT_REDACTED")],
                    "TXT_REDACTED": [],
                    "TXT_REDACTED": "TXT_REDACTED",
                    "TXT_REDACTED": "TXT_REDACTED"                                ,
                },
            )
            if label not in group["TXT_REDACTED"]:
                group["TXT_REDACTED"].append(label)

    derived_entries: List[Dict[str, Any]] = []
    for rule_id, rule in sorted((derivation_rules.get("TXT_REDACTED") or {}).items()):
        label = str(rule.get("TXT_REDACTED") or rule_id)
        aliases = [str(item) for item in (rule.get("TXT_REDACTED") or []) if str(item).strip()]
        operands = [str(item) for item in (rule.get("TXT_REDACTED") or []) if str(item).strip()]
        section_num = _resolve_section_num(operands[1] if operands else label, fact_section_index)
        entry = {
            "TXT_REDACTED": "TXT_REDACTED"                                                    ,
            "TXT_REDACTED": "TXT_REDACTED",
            "TXT_REDACTED": section_num,
            "TXT_REDACTED": label,
            "TXT_REDACTED": [label, *aliases],
            "TXT_REDACTED": "TXT_REDACTED",
            "TXT_REDACTED": "TXT_REDACTED"                           ,
            "TXT_REDACTED": _group_key(label),
            "TXT_REDACTED": rule_id,
            "TXT_REDACTED": str(rule.get("TXT_REDACTED") or rule.get("TXT_REDACTED") or "TXT_REDACTED"),
            "TXT_REDACTED": operands,
        }
        derived_entries.append(entry)
        if section_num and section_num in sections:
            sections[section_num]["TXT_REDACTED"].setdefault(label, entry)

    flat_entries: List[Dict[str, Any]] = []
    tree_sections: List[Dict[str, Any]] = []
    for section_num in sorted(sections):
        payload = sections[section_num]
        field_entries = sorted(payload["TXT_REDACTED"].values(), key=lambda item: item["TXT_REDACTED"])
        group_entries = sorted(payload["TXT_REDACTED"].values(), key=lambda item: item["TXT_REDACTED"])
        score_entries = sorted(payload["TXT_REDACTED"], key=lambda item: item["TXT_REDACTED"])
        flat_entries.extend(score_entries)
        flat_entries.extend(group_entries)
        flat_entries.extend(field_entries)
        flat_entries.extend(metric_groups.get(section_num, {}).values())
        tree_sections.append(
            {
                "TXT_REDACTED": section_num,
                "TXT_REDACTED": payload["TXT_REDACTED"],
                "TXT_REDACTED": payload["TXT_REDACTED"],
                "TXT_REDACTED": score_entries,
                "TXT_REDACTED": group_entries,
                "TXT_REDACTED": field_entries,
                "TXT_REDACTED": sorted(metric_groups.get(section_num, {}).values(), key=lambda item: item["TXT_REDACTED"]),
            }
        )

    flat_entries.extend(derived_entries)
    version_seed = json.dumps(
        {
            "TXT_REDACTED": tree_sections,
            "TXT_REDACTED": derived_entries,
            "TXT_REDACTED": derivation_rules.get("TXT_REDACTED", "TXT_REDACTED"),
        },
        ensure_ascii=False,
        sort_keys=True,
    )
    version = hashlib.sha256(version_seed.encode("TXT_REDACTED")).hexdigest()[:2]
    return {
        "TXT_REDACTED": version,
        "TXT_REDACTED": dt.datetime.now(dt.timezone.utc).isoformat(),
        "TXT_REDACTED": str(derivation_rules.get("TXT_REDACTED") or "TXT_REDACTED"),
        "TXT_REDACTED": tree_sections,
        "TXT_REDACTED": derived_entries,
        "TXT_REDACTED": sorted(flat_entries, key=lambda item: item["TXT_REDACTED"]),
    }


def build_output_schema_markdown(payload: Dict[str, Any]) -> str:
    lines = [
        "TXT_REDACTED",
        "TXT_REDACTED",
        "TXT_REDACTED"                                                  ,
        "TXT_REDACTED"                                                                             ,
        "TXT_REDACTED",
    ]
    for section in payload.get("TXT_REDACTED", []):
        lines.append("TXT_REDACTED"                          )
        score_labels = [item.get("TXT_REDACTED") for item in (section.get("TXT_REDACTED") or []) if item.get("TXT_REDACTED")]
        if score_labels:
            lines.append("TXT_REDACTED"                                )
        group_labels = [item.get("TXT_REDACTED") for item in (section.get("TXT_REDACTED") or []) if item.get("TXT_REDACTED")]
        if group_labels:
            lines.append("TXT_REDACTED"                                     )
        field_labels = [item.get("TXT_REDACTED") for item in (section.get("TXT_REDACTED") or []) if item.get("TXT_REDACTED")]
        if field_labels:
            lines.append("TXT_REDACTED"                                     )
        metric_labels = [item.get("TXT_REDACTED") for item in (section.get("TXT_REDACTED") or []) if item.get("TXT_REDACTED")]
        if metric_labels:
            lines.append("TXT_REDACTED"                                          )
        lines.append("TXT_REDACTED")
    if payload.get("TXT_REDACTED"):
        lines.append("TXT_REDACTED")
        for entry in payload.get("TXT_REDACTED", [])[:3]:
            formula = str(entry.get("TXT_REDACTED") or "TXT_REDACTED")
            lines.append("TXT_REDACTED"                                  )
        lines.append("TXT_REDACTED")
    return "TXT_REDACTED".join(lines).strip() + "TXT_REDACTED"
