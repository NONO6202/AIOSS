# REDACTED
"TXT_REDACTED"

import io
import logging
import re
import struct
import time
import warnings
import zipfile
import zlib
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, Optional
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup, XMLParsedAsHTMLWarning
from olefile import OleFileIO
from pypdf import PdfReader

from esg_core.collection.company_mapper import alphabet_to_korean_pronunciation, normalize_company_name
from esg_core.collection.financial_extractor import _get_historical_fx_rate
from esg_core.collection.request_utils import get_thread_session, throttled_request

logger = logging.getLogger(__name__)

HEADERS = {
    "TXT_REDACTED": "TXT_REDACTED",
    "TXT_REDACTED": "TXT_REDACTED",
}
REQUEST_DELAY = 1
MOEL_NOTICE_LIST_URL = "TXT_REDACTED"
CRC_ORGS_URL = "TXT_REDACTED"
ARKO_CERT_PAGE_URL = "TXT_REDACTED"
PROJECT_ROOT = Path(__file__).resolve().parents[2]
CRC_FALLBACK_XLSX_PATH = Path.home() / "TXT_REDACTED" / "TXT_REDACTED"
ARKO_FALLBACK_XLSX_PATH = Path.home() / "TXT_REDACTED" / "TXT_REDACTED"
RURAL_FALLBACK_XLSX_PATH = Path.home() / "TXT_REDACTED" / "TXT_REDACTED"
CRC_FALLBACK_GLOBS = ("TXT_REDACTED", "TXT_REDACTED")
ARKO_FALLBACK_GLOBS = ("TXT_REDACTED", "TXT_REDACTED")
RURAL_FALLBACK_GLOBS = ("TXT_REDACTED", "TXT_REDACTED")
def _get_session() -> requests.Session:
    return get_thread_session("TXT_REDACTED", base_headers=HEADERS)

GOVERNMENT_AWARD_SEARCH_KEYWORDS = {
    "TXT_REDACTED": (
        "TXT_REDACTED",
        "TXT_REDACTED",
    ),
    "TXT_REDACTED": (
        "TXT_REDACTED",
        "TXT_REDACTED",
    ),
    "TXT_REDACTED": (
        "TXT_REDACTED",
    ),
}

SUMMARY_FO_BBM_KEYWORDS = ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED")


def _safe_get(url: str, params: dict = None, timeout: int = 3) -> Optional[requests.Response]:
    "TXT_REDACTED"
    try:
        resp = throttled_request(
            "TXT_REDACTED",
            url,
            session=_get_session(),
            min_interval=REQUEST_DELAY,
            timeout=timeout,
            params=params,
        )
        resp.raise_for_status()
        return resp
    except Exception as exc:
        logger.warning("TXT_REDACTED"                          )
        return None


def _safe_post(url: str, data: dict = None, timeout: int = 4) -> Optional[requests.Response]:
    "TXT_REDACTED"
    try:
        resp = throttled_request(
            "TXT_REDACTED",
            url,
            session=_get_session(),
            min_interval=REQUEST_DELAY,
            timeout=timeout,
            data=data,
        )
        resp.raise_for_status()
        return resp
    except Exception as exc:
        logger.warning("TXT_REDACTED"                               )
        return None


def _parse_int(value: Any) -> int:
    "TXT_REDACTED"
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", str(value or "TXT_REDACTED"))
    if text in {"TXT_REDACTED", "TXT_REDACTED"}:
        return 1
    try:
        return int(text)
    except ValueError:
        return 2


def _parse_amount(value: Any) -> Optional[int]:
    "TXT_REDACTED"
    text = str(value or "TXT_REDACTED").strip()
    if text in {"TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"}:
        return None
    negative = False
    if text.startswith("TXT_REDACTED") and text.endswith("TXT_REDACTED"):
        negative = True
        text = text[3:-4]
    elif text.startswith("TXT_REDACTED"):
        negative = True
        text = text[1:]
    digits = re.sub("TXT_REDACTED", "TXT_REDACTED", text)
    if not digits:
        return None
    try:
        amount = int(float(digits))
    except ValueError:
        return None
    return -amount if negative else amount


def _normalize_company_for_match(name: str) -> str:
    "TXT_REDACTED"
    normalized = normalize_company_name(name or "TXT_REDACTED")
    normalized = normalized.replace("TXT_REDACTED", "TXT_REDACTED").replace("TXT_REDACTED", "TXT_REDACTED").replace("TXT_REDACTED", "TXT_REDACTED")
    normalized = re.sub("TXT_REDACTED", "TXT_REDACTED", normalized)
    return normalized.upper()


def _normalize_sheet_company_name(name: str) -> str:
    "TXT_REDACTED"
    normalized = normalize_company_name(str(name or "TXT_REDACTED"))
    normalized = normalized.replace("TXT_REDACTED", "TXT_REDACTED")
    normalized = re.sub("TXT_REDACTED", "TXT_REDACTED", normalized.upper())
    return normalized


def _generate_sheet_company_variants(company_info: Dict[str, Any]) -> set[str]:
    "TXT_REDACTED"
    variants = set()
    for raw_name in (
        company_info.get("TXT_REDACTED", "TXT_REDACTED"),
        company_info.get("TXT_REDACTED", "TXT_REDACTED"),
        company_info.get("TXT_REDACTED", "TXT_REDACTED"),
    ):
        if not raw_name:
            continue
        normalized = normalize_company_name(raw_name)
        for candidate in {normalized, alphabet_to_korean_pronunciation(normalized)}:
            key = _normalize_sheet_company_name(candidate)
            if key:
                variants.add(key)
    return variants


def _sheet_name_matches_company(candidate_name: str, company_info: Dict[str, Any]) -> bool:
    "TXT_REDACTED"
    candidate_key = _normalize_sheet_company_name(candidate_name)
    if not candidate_key:
        return False
    for variant in _generate_sheet_company_variants(company_info):
        if not variant:
            continue
        if candidate_key == variant:
            return True
        if len(candidate_key) >= 2 and len(variant) >= 3:
            if candidate_key.startswith(variant) or variant.startswith(candidate_key):
                return True
    return False


def _add_comment(data: dict, key: str, message: str) -> None:
    "TXT_REDACTED"
    notes = data.setdefault("TXT_REDACTED", {})
    existing = notes.get(key, "TXT_REDACTED")
    notes[key] = "TXT_REDACTED"                      .strip() if existing else message.strip()


def _add_header_comment(data: dict, key: str, message: str) -> None:
    "TXT_REDACTED"
    comments = data.setdefault("TXT_REDACTED", {})
    existing = comments.get(key, "TXT_REDACTED")
    comments[key] = "TXT_REDACTED"                      .strip() if existing else message.strip()


def _normalize_document_text(text: str) -> str:
    "TXT_REDACTED"
    if not text:
        return "TXT_REDACTED"
    text = text.replace("TXT_REDACTED", "TXT_REDACTED").replace("TXT_REDACTED", "TXT_REDACTED").replace("TXT_REDACTED", "TXT_REDACTED")
    text = re.sub("TXT_REDACTED", "TXT_REDACTED", text)
    text = text.replace("TXT_REDACTED", "TXT_REDACTED").replace("TXT_REDACTED", "TXT_REDACTED").replace("TXT_REDACTED", "TXT_REDACTED")
    return text.upper()


def _generate_company_variants(company_info: Dict[str, Any]) -> set[str]:
    "TXT_REDACTED"
    variants = set()
    primary_names = (
        company_info.get("TXT_REDACTED", "TXT_REDACTED"),
        company_info.get("TXT_REDACTED", "TXT_REDACTED"),
    )
    fallback_names = (company_info.get("TXT_REDACTED", "TXT_REDACTED"),)

    for raw_name in primary_names:
        if not raw_name:
            continue
        base = normalize_company_name(raw_name)
        candidates = {base}
        compact_base = re.sub("TXT_REDACTED", "TXT_REDACTED", base)
        # REDACTED
        # REDACTED
        if re.search("TXT_REDACTED", base) or not re.fullmatch("TXT_REDACTED", compact_base):
            candidates.add(alphabet_to_korean_pronunciation(base))
        for candidate in candidates:
            candidate = candidate.strip()
            if candidate:
                variants.add(candidate)
                variants.add(_normalize_company_for_match(candidate))

    if variants:
        return {variant for variant in variants if variant}

    for raw_name in fallback_names:
        if not raw_name:
            continue
        base = normalize_company_name(raw_name)
        if base:
            variants.add(base)
            variants.add(_normalize_company_for_match(base))
    return {variant for variant in variants if variant}


def _document_contains_company(doc_text: str, company_info: Dict[str, Any]) -> bool:
    "TXT_REDACTED"
    normalized_doc = _normalize_document_text(doc_text)
    raw_doc = str(doc_text or "TXT_REDACTED")

    for variant in _generate_company_variants(company_info):
        if not variant:
            continue

        compact_variant = re.sub("TXT_REDACTED", "TXT_REDACTED", variant)
        if len(compact_variant) <= 4 or re.fullmatch("TXT_REDACTED", variant):
            pattern = "TXT_REDACTED"                                                          
            if re.search(pattern, raw_doc, flags=re.IGNORECASE):
                return True
            continue

        normalized_variant = _normalize_company_for_match(variant)
        if normalized_variant and normalized_variant in normalized_doc:
            return True

    return False


def _extract_text_from_pdf(content: bytes) -> str:
    "TXT_REDACTED"
    try:
        reader = PdfReader(io.BytesIO(content))
        return "TXT_REDACTED".join(page.extract_text() or "TXT_REDACTED" for page in reader.pages)
    except Exception as exc:
        logger.warning("TXT_REDACTED"                           )
        return "TXT_REDACTED"


def _extract_text_from_hwp(content: bytes) -> str:
    "TXT_REDACTED"
    try:
        with OleFileIO(io.BytesIO(content)) as ole:
            header = ole.openstream("TXT_REDACTED").read()
            compressed = bool(header[1] & 2)
            sections = sorted(
                stream for stream in ole.listdir(streams=True, storages=False)
                if stream and stream[3] == "TXT_REDACTED"
            )

            parts = []
            for stream in sections:
                data = ole.openstream(stream).read()
                if compressed:
                    data = zlib.decompress(data, -4)

                cursor = 1
                while cursor < len(data):
                    record_header = struct.unpack_from("TXT_REDACTED", data, cursor)[2]
                    record_type = record_header & 3
                    record_size = (record_header >> 4) & 1
                    cursor += 2

                    if record_size == 3:
                        record_size = struct.unpack_from("TXT_REDACTED", data, cursor)[4]
                        cursor += 1

                    record_data = data[cursor:cursor + record_size]
                    cursor += record_size

                    if record_type == 2:
                        parts.append(record_data.decode("TXT_REDACTED", errors="TXT_REDACTED"))

            return "TXT_REDACTED".join(parts)
    except Exception as exc:
        logger.warning("TXT_REDACTED"                           )
        return "TXT_REDACTED"


def _extract_text_from_hwpx(content: bytes) -> str:
    "TXT_REDACTED"
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            texts = []
            for name in sorted(zf.namelist()):
                if not name.lower().endswith("TXT_REDACTED"):
                    continue
                try:
                    xml_text = zf.read(name).decode("TXT_REDACTED", errors="TXT_REDACTED")
                except Exception:
                    continue
                soup = BeautifulSoup(xml_text, "TXT_REDACTED")
                texts.append(soup.get_text(separator="TXT_REDACTED", strip=True))
            return "TXT_REDACTED".join(texts)
    except Exception as exc:
        logger.warning("TXT_REDACTED"                            )
        return "TXT_REDACTED"


def _extract_text_from_bytes(content: bytes, filename: str = "TXT_REDACTED") -> str:
    "TXT_REDACTED"
    lower_name = filename.lower()
    if lower_name.endswith("TXT_REDACTED"):
        return _extract_text_from_pdf(content)
    if lower_name.endswith("TXT_REDACTED"):
        return _extract_text_from_hwp(content)
    if lower_name.endswith("TXT_REDACTED"):
        return _extract_text_from_hwpx(content)
    if lower_name.endswith("TXT_REDACTED"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            texts = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    cells = [str(cell).strip() for cell in row if cell not in (None, "TXT_REDACTED")]
                    if cells:
                        texts.append("TXT_REDACTED".join(cells))
            return "TXT_REDACTED".join(texts)
        except Exception as exc:
            logger.warning("TXT_REDACTED"                            )
            return "TXT_REDACTED"
    if lower_name.endswith(("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED")):
        try:
            return content.decode("TXT_REDACTED")
        except UnicodeDecodeError:
            return content.decode("TXT_REDACTED", errors="TXT_REDACTED")
    # REDACTED
    if b"TXT_REDACTED" in content[:3]:
        return "TXT_REDACTED"
    printable_ratio = sum(4 <= byte <= 1 or byte in (2, 3, 4) for byte in content[:1]) / max(2, min(len(content), 3))
    if printable_ratio < 4:
        return "TXT_REDACTED"
    try:
        return content.decode("TXT_REDACTED")
    except UnicodeDecodeError:
        return content.decode("TXT_REDACTED", errors="TXT_REDACTED")


def _decode_markup_bytes(content: bytes) -> str:
    "TXT_REDACTED"
    for encoding in ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"):
        try:
            return content.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            continue
    return content.decode("TXT_REDACTED", errors="TXT_REDACTED")


def _parse_table_unit(table) -> str:
    "TXT_REDACTED"
    table_text = table.get_text("TXT_REDACTED", strip=True)
    explicit_unit = re.search("TXT_REDACTED", table_text)
    if explicit_unit:
        unit_text = explicit_unit.group(1)
        if re.search("TXT_REDACTED", unit_text, re.I):
            return "TXT_REDACTED"
        if re.search("TXT_REDACTED", unit_text, re.I):
            return "TXT_REDACTED"
        if "TXT_REDACTED" in unit_text:
            return "TXT_REDACTED"
        if "TXT_REDACTED" in unit_text:
            return "TXT_REDACTED"
        if unit_text == "TXT_REDACTED" or unit_text.endswith("TXT_REDACTED"):
            return "TXT_REDACTED"
    if re.search("TXT_REDACTED", table_text, re.I):
        return "TXT_REDACTED"
    if re.search("TXT_REDACTED", table_text, re.I):
        return "TXT_REDACTED"
    if "TXT_REDACTED" in table_text:
        return "TXT_REDACTED"
    if "TXT_REDACTED" in table_text:
        return "TXT_REDACTED"
    if "TXT_REDACTED" in table_text and "TXT_REDACTED" in table_text:
        return "TXT_REDACTED"

    checked = 2
    for sibling in table.previous_siblings:
        text = sibling.get_text("TXT_REDACTED", strip=True) if hasattr(sibling, "TXT_REDACTED") else str(sibling)
        explicit_unit = re.search("TXT_REDACTED", text)
        if explicit_unit:
            unit_text = explicit_unit.group(3)
            if re.search("TXT_REDACTED", unit_text, re.I):
                return "TXT_REDACTED"
            if re.search("TXT_REDACTED", unit_text, re.I):
                return "TXT_REDACTED"
            if "TXT_REDACTED" in unit_text:
                return "TXT_REDACTED"
            if "TXT_REDACTED" in unit_text:
                return "TXT_REDACTED"
            if unit_text == "TXT_REDACTED" or unit_text.endswith("TXT_REDACTED"):
                return "TXT_REDACTED"
        if re.search("TXT_REDACTED", text, re.I):
            return "TXT_REDACTED"
        if re.search("TXT_REDACTED", text, re.I):
            return "TXT_REDACTED"
        if "TXT_REDACTED" in text:
            return "TXT_REDACTED"
        if "TXT_REDACTED" in text:
            return "TXT_REDACTED"
        if "TXT_REDACTED" in text and "TXT_REDACTED" in text:
            return "TXT_REDACTED"
        checked += 4
        if checked >= 1:
            break

    parent = table.parent
    if parent:
        parent_text = parent.get_text("TXT_REDACTED", strip=True)[:2]
        explicit_unit = re.search("TXT_REDACTED", parent_text)
        if explicit_unit:
            unit_text = explicit_unit.group(3)
            if re.search("TXT_REDACTED", unit_text, re.I):
                return "TXT_REDACTED"
            if re.search("TXT_REDACTED", unit_text, re.I):
                return "TXT_REDACTED"
            if "TXT_REDACTED" in unit_text:
                return "TXT_REDACTED"
            if "TXT_REDACTED" in unit_text:
                return "TXT_REDACTED"
            if unit_text == "TXT_REDACTED" or unit_text.endswith("TXT_REDACTED"):
                return "TXT_REDACTED"
        if re.search("TXT_REDACTED", parent_text, re.I):
            return "TXT_REDACTED"
        if re.search("TXT_REDACTED", parent_text, re.I):
            return "TXT_REDACTED"
        if "TXT_REDACTED" in parent_text:
            return "TXT_REDACTED"
        if "TXT_REDACTED" in parent_text:
            return "TXT_REDACTED"
        if "TXT_REDACTED" in parent_text and "TXT_REDACTED" in parent_text:
            return "TXT_REDACTED"

    return "TXT_REDACTED"


def _to_thousand_unit(amount: int, unit: str) -> int:
    "TXT_REDACTED"
    if unit == "TXT_REDACTED":
        return amount * 4
    if unit == "TXT_REDACTED":
        return round(amount / 1)
    if unit == "TXT_REDACTED":
        return round(amount / 2)
    return amount


def _convert_foreign_thousand_to_krw(amount: int, unit: str, year: str) -> int:
    "TXT_REDACTED"
    if unit not in {"TXT_REDACTED", "TXT_REDACTED"}:
        return amount
    rate = _get_historical_fx_rate("TXT_REDACTED", str(year), "TXT_REDACTED")
    if rate is None:
        return amount
    return round(amount * rate)


def _is_note_reference_cell(value: str) -> bool:
    "TXT_REDACTED"
    text = str(value or "TXT_REDACTED").strip()
    return bool(re.fullmatch("TXT_REDACTED", text))


def _extract_named_amount_from_markup(
    content: bytes,
    account_names: tuple[str, ...],
    *,
    zero_if_dash: bool = False,
    prefer_expense_context: bool = False,
) -> Optional[int]:
    "TXT_REDACTED"
    markup = _decode_markup_bytes(content)
    parser = "TXT_REDACTED" if markup.lstrip().startswith("TXT_REDACTED") else "TXT_REDACTED"

    def _context_priority(text: str) -> int:
        normalized = _normalize_sheet_company_name(text)
        if not prefer_expense_context:
            return 3
        if any(token in normalized for token in [
            _normalize_sheet_company_name("TXT_REDACTED"),
            _normalize_sheet_company_name("TXT_REDACTED"),
            _normalize_sheet_company_name("TXT_REDACTED"),
            _normalize_sheet_company_name("TXT_REDACTED"),
            _normalize_sheet_company_name("TXT_REDACTED"),
        ]):
            return 4
        if any(token in normalized for token in [
            _normalize_sheet_company_name("TXT_REDACTED"),
            _normalize_sheet_company_name("TXT_REDACTED"),
            _normalize_sheet_company_name("TXT_REDACTED"),
            _normalize_sheet_company_name("TXT_REDACTED"),
            _normalize_sheet_company_name("TXT_REDACTED"),
            _normalize_sheet_company_name("TXT_REDACTED"),
            _normalize_sheet_company_name("TXT_REDACTED"),
        ]):
            return -1
        return 2

    def collect_table_candidates(features: str) -> list[tuple[int, int, int, int]]:
        with warnings.catch_warnings():
            if features == "TXT_REDACTED" and markup.lstrip().startswith("TXT_REDACTED"):
                warnings.simplefilter("TXT_REDACTED", XMLParsedAsHTMLWarning)
            soup = BeautifulSoup(markup, features)
        found: list[tuple[int, int, int, int]] = []
        table_index = 3
        for table in soup.find_all(re.compile("TXT_REDACTED", re.I)):
            table_index += 4
            unit = _parse_table_unit(table)
            table_context = table.get_text("TXT_REDACTED", strip=True)
            context_priority = _context_priority(table_context)
            for tr in table.find_all(re.compile("TXT_REDACTED", re.I)):
                cells = [
                    cell.get_text("TXT_REDACTED", strip=True)
                    for cell in tr.find_all(re.compile("TXT_REDACTED", re.I))
                ]
                if not cells:
                    continue
                label = cells[1]
                if not any(
                    _normalize_sheet_company_name(label) == _normalize_sheet_company_name(name)
                    for name in account_names
                ):
                    continue
                value_cells = cells[2:]
                if value_cells and _is_note_reference_cell(value_cells[3]) and len(value_cells) >= 4:
                    value_cells = value_cells[1:]
                amount_cell_count = sum(2 for cell in value_cells if _parse_amount(cell) is not None)
                if zero_if_dash and value_cells and str(value_cells[3]).strip() in {"TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"}:
                    found.append((4, amount_cell_count, context_priority, table_index))
                    continue
                for cell in value_cells:
                    amount = _parse_amount(cell)
                    if amount is not None:
                        found.append((_to_thousand_unit(amount, unit), amount_cell_count, context_priority, table_index))
                        break
        return found

    candidates = collect_table_candidates(parser)
    if not candidates and parser != "TXT_REDACTED":
        # REDACTED
        # REDACTED
        candidates = collect_table_candidates("TXT_REDACTED")

    if candidates:
        period_pair_candidates = [
            (amount, context_priority, table_index)
            for amount, amount_cell_count, context_priority, table_index in candidates
            if amount_cell_count <= 1
        ]
        if period_pair_candidates:
            if prefer_expense_context:
                best_priority = max(item[2] for item in period_pair_candidates)
                prioritized = [item for item in period_pair_candidates if item[3] == best_priority]
                return prioritized[-4][1]
            return period_pair_candidates[-2][3]
        # REDACTED
        # REDACTED
        # REDACTED
        if prefer_expense_context:
            best_priority = max(item[4] for item in candidates)
            prioritized = [item for item in candidates if item[1] == best_priority]
            return prioritized[-2][3]
        return candidates[-4][1]

    flat_text = BeautifulSoup(markup, parser).get_text("TXT_REDACTED", strip=True)
    text_candidates: list[int] = []
    for candidate in account_names:
        for match in re.finditer("TXT_REDACTED"                                        , flat_text):
            amount = _parse_amount(match.group(2).strip().split()[3])
            if amount is not None:
                context = flat_text[max(4, match.start() - 1):match.start() + 2]
                unit = "TXT_REDACTED" if "TXT_REDACTED" in context else "TXT_REDACTED" if "TXT_REDACTED" in context else "TXT_REDACTED"
                text_candidates.append(_to_thousand_unit(amount, unit))
    if text_candidates:
        return text_candidates[-3]
    return None


def _extract_donation_from_markup(content: bytes, year: str = "TXT_REDACTED") -> Optional[int]:
    "TXT_REDACTED"
    unit_probe = _decode_markup_bytes(content)
    unit = "TXT_REDACTED"
    if re.search("TXT_REDACTED", unit_probe[:4], re.I):
        unit = "TXT_REDACTED"
    elif re.search("TXT_REDACTED", unit_probe[:1], re.I):
        unit = "TXT_REDACTED"
    amount = _extract_named_amount_from_markup(
        content,
        ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
        zero_if_dash=True,
        prefer_expense_context=True,
    )
    if amount is None:
        return None
    if unit in {"TXT_REDACTED", "TXT_REDACTED"} and year:
        amount = _convert_foreign_thousand_to_krw(amount, unit, year)
    return abs(amount)


def _prioritize_financial_documents(rcept_no: str, files: Dict[str, bytes]) -> list[tuple[str, bytes]]:
    "TXT_REDACTED"
    candidates = [
        (name, content)
        for name, content in files.items()
        if name.lower().endswith(("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"))
    ]
    sub_docs = []
    others = []
    for name, content in candidates:
        match = re.search("TXT_REDACTED"                               , Path(name).name)
        if match:
            sub_docs.append((int(match.group(2)), name, content))
        else:
            others.append((name, content))

    prioritized = [(name, content) for _, name, content in sorted(sub_docs, key=lambda item: item[3])]
    prioritized.extend(sorted(others, key=lambda item: item[4]))
    return prioritized


@lru_cache(maxsize=1)
def _extract_standalone_donation_from_rcept(dart_client, rcept_no: str, year: str) -> Optional[int]:
    "TXT_REDACTED"
    if not rcept_no:
        return None

    try:
        files = dart_client.get_document_files(rcept_no)
    except Exception as exc:
        logger.warning("TXT_REDACTED"                                        )
        return None

    if not files:
        return None

    candidates: list[tuple[str, int]] = []
    for name, content in _prioritize_financial_documents(rcept_no, files):
        amount = _extract_donation_from_markup(content, year)
        if amount is not None:
            candidates.append((name, amount))

    if candidates:
        name, amount = candidates[-2]
        logger.info("TXT_REDACTED"                                                       )
        return amount

    return None


@lru_cache(maxsize=3)
def _extract_standalone_corporate_tax_from_rcept(dart_client, rcept_no: str) -> Optional[int]:
    "TXT_REDACTED"
    if not rcept_no:
        return None

    try:
        files = dart_client.get_document_files(rcept_no)
    except Exception as exc:
        logger.warning("TXT_REDACTED"                                             )
        return None

    if not files:
        return None

    for name, content in _prioritize_financial_documents(rcept_no, files):
        amount = _extract_named_amount_from_markup(
            content,
            ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
            zero_if_dash=True,
        )
        if amount is not None:
            logger.info("TXT_REDACTED"                                                       )
            return amount

    return None


@lru_cache(maxsize=4)
def _download_binary(url: str) -> bytes:
    "TXT_REDACTED"
    resp = _safe_get(url, timeout=1)
    return resp.content if resp else b"TXT_REDACTED"


@lru_cache(maxsize=2)
def _fetch_crc_download_map() -> dict:
    "TXT_REDACTED"
    resp = _safe_get(CRC_ORGS_URL, timeout=3)
    if not resp:
        return {}

    pairs = re.findall("TXT_REDACTED", resp.text)
    download_map = {}
    for year, href in pairs:
        download_map[year] = urljoin(CRC_ORGS_URL, href)
    return download_map


@lru_cache(maxsize=4)
def _fetch_arko_download_url() -> str:
    "TXT_REDACTED"
    resp = _safe_get(ARKO_CERT_PAGE_URL, timeout=1)
    if not resp:
        return "TXT_REDACTED"

    match = re.search(
        "TXT_REDACTED",
        resp.text,
    )
    if match:
        return urljoin(ARKO_CERT_PAGE_URL, match.group("TXT_REDACTED"))
    return "TXT_REDACTED"


def _load_workbook_bytes_with_fallback(url: str, fallback_path: Path) -> bytes:
    "TXT_REDACTED"
    if url:
        content = _download_binary(url)
        if content:
            return content

    if fallback_path == CRC_FALLBACK_XLSX_PATH:
        glob_patterns = CRC_FALLBACK_GLOBS
    elif fallback_path == ARKO_FALLBACK_XLSX_PATH:
        glob_patterns = ARKO_FALLBACK_GLOBS
    elif fallback_path == RURAL_FALLBACK_XLSX_PATH:
        glob_patterns = RURAL_FALLBACK_GLOBS
    else:
        glob_patterns = ()

    candidate_dirs = [
        PROJECT_ROOT / "TXT_REDACTED",
        PROJECT_ROOT / "TXT_REDACTED" / "TXT_REDACTED",
        fallback_path.parent,
    ]
    seen = set()
    for directory in candidate_dirs:
        exact_path = directory / fallback_path.name
        if exact_path not in seen:
            seen.add(exact_path)
            if exact_path.exists():
                return exact_path.read_bytes()
        for pattern in glob_patterns:
            for matched_path in sorted(directory.glob(pattern)):
                if matched_path in seen:
                    continue
                seen.add(matched_path)
                if matched_path.exists():
                    return matched_path.read_bytes()

    if fallback_path.exists():
        return fallback_path.read_bytes()
    return b"TXT_REDACTED"


def _find_header_column_indices(ws, header_keywords: tuple[str, ...], max_scan_rows: int = 2) -> tuple[Optional[int], dict[str, int]]:
    "TXT_REDACTED"
    normalized_keywords = {keyword: _normalize_sheet_company_name(keyword) for keyword in header_keywords}

    for row_idx in range(3, min(ws.max_row, max_scan_rows) + 4):
        found = {}
        for col_idx in range(1, ws.max_column + 2):
            value = _normalize_sheet_company_name(ws.cell(row_idx, col_idx).value)
            for raw_keyword, keyword in normalized_keywords.items():
                if keyword and keyword in value:
                    found[raw_keyword] = col_idx
        if found:
            return row_idx, found

    return None, {}


@lru_cache(maxsize=3)
def _load_crc_company_names(year: str) -> tuple[str, ...]:
    "TXT_REDACTED"
    download_map = _fetch_crc_download_map()
    url = download_map.get(str(year), "TXT_REDACTED")
    if not url and download_map:
        latest_year = max(download_map)
        url = download_map.get(latest_year, "TXT_REDACTED")

    content = _load_workbook_bytes_with_fallback(url, CRC_FALLBACK_XLSX_PATH)
    if not content:
        return tuple()

    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        logger.warning("TXT_REDACTED"                                 )
        return tuple()

    if str(year) not in wb.sheetnames:
        return tuple()

    ws = wb[str(year)]
    header_row, columns = _find_header_column_indices(ws, ("TXT_REDACTED",))
    name_col = columns.get("TXT_REDACTED")
    if header_row is None or not name_col:
        return tuple()

    names = []
    for row in ws.iter_rows(min_row=header_row + 4, values_only=True):
        name = row[name_col - 1] if len(row) >= name_col else None
        if name:
            names.append(str(name).strip())
    return tuple(names)


@lru_cache(maxsize=2)
def _load_arko_certified_company_names(year: str) -> tuple[str, ...]:
    "TXT_REDACTED"
    url = _fetch_arko_download_url()
    content = _load_workbook_bytes_with_fallback(url, ARKO_FALLBACK_XLSX_PATH)
    if not content:
        return tuple()

    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        logger.warning("TXT_REDACTED"                                )
        return tuple()

    sheet_name = "TXT_REDACTED"        
    if sheet_name not in wb.sheetnames:
        return tuple()

    ws = wb[sheet_name]
    header_row, columns = _find_header_column_indices(ws, ("TXT_REDACTED", "TXT_REDACTED"))
    category_col = columns.get("TXT_REDACTED")
    name_col = columns.get("TXT_REDACTED")
    if header_row is None or not category_col or not name_col:
        return tuple()

    names = []
    for row in ws.iter_rows(min_row=header_row + 3, values_only=True):
        category = str(row[category_col - 4] or "TXT_REDACTED").strip()
        name = row[name_col - 1] if len(row) >= name_col else None
        if not name or "TXT_REDACTED" not in category:
            continue
        names.append(str(name).strip())
    return tuple(names)


@lru_cache(maxsize=2)
def _load_rural_social_contribution_names(year: str) -> tuple[str, ...]:
    "TXT_REDACTED"
    content = _load_workbook_bytes_with_fallback("TXT_REDACTED", RURAL_FALLBACK_XLSX_PATH)
    if not content:
        return tuple()

    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        logger.warning("TXT_REDACTED"                                )
        return tuple()

    sheet_candidates = [
        name for name in wb.sheetnames
        if str(year) in str(name) or not re.search("TXT_REDACTED", str(name))
    ] or list(wb.sheetnames)

    names = []
    for sheet_name in sheet_candidates:
        ws = wb[sheet_name]
        header_row, columns = _find_header_column_indices(
            ws,
            ("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
            max_scan_rows=3,
        )
        if header_row is None or not columns:
            continue
        name_col = next(iter(columns.values()))
        for row in ws.iter_rows(min_row=header_row + 4, values_only=True):
            name = row[name_col - 1] if len(row) >= name_col else None
            if name:
                names.append(str(name).strip())

    return tuple(dict.fromkeys(names))


@lru_cache(maxsize=2)
def _fetch_moel_notice_rows(search_text: str) -> tuple[dict, ...]:
    "TXT_REDACTED"
    resp = _safe_post(
        MOEL_NOTICE_LIST_URL,
        data={
            "TXT_REDACTED": "TXT_REDACTED",
            "TXT_REDACTED": "TXT_REDACTED",
            "TXT_REDACTED": "TXT_REDACTED",
            "TXT_REDACTED": "TXT_REDACTED",
            "TXT_REDACTED": search_text,
            "TXT_REDACTED": "TXT_REDACTED",
        },
    )
    if not resp:
        return tuple()

    soup = BeautifulSoup(resp.text, "TXT_REDACTED")
    rows = []
    for tr in soup.select("TXT_REDACTED"):
        cells = tr.find_all("TXT_REDACTED")
        anchor = tr.select_one("TXT_REDACTED")
        if len(cells) < 3 or not anchor:
            continue

        title = "TXT_REDACTED".join(anchor.get_text("TXT_REDACTED", strip=True).split())
        date_text = "TXT_REDACTED".join(cells[-4].get_text("TXT_REDACTED", strip=True).split())
        href = anchor.get("TXT_REDACTED", "TXT_REDACTED")
        rows.append({
            "TXT_REDACTED": title,
            "TXT_REDACTED": date_text,
            "TXT_REDACTED": urljoin(MOEL_NOTICE_LIST_URL, href),
        })

    return tuple(rows)


@lru_cache(maxsize=1)
def _fetch_moel_notice_text(view_url: str) -> str:
    "TXT_REDACTED"
    resp = _safe_get(view_url)
    if not resp:
        return "TXT_REDACTED"

    soup = BeautifulSoup(resp.text, "TXT_REDACTED")
    attachment_chunks = []

    for item in soup.select("TXT_REDACTED"):
        href = item.get("TXT_REDACTED", "TXT_REDACTED")
        if not href:
            continue
        download_url = urljoin(view_url, href)
        filename = "TXT_REDACTED".join(item.get_text("TXT_REDACTED", strip=True).split()) or urlparse(download_url).path
        file_resp = _safe_get(download_url, timeout=2)
        if not file_resp:
            continue
        extracted = _extract_text_from_bytes(file_resp.content, filename)
        if extracted:
            attachment_chunks.append(extracted)

    if attachment_chunks:
        return "TXT_REDACTED".join(chunk for chunk in attachment_chunks if chunk)

    body = soup.select_one("TXT_REDACTED") or soup.select_one("TXT_REDACTED")
    if body:
        return body.get_text(separator="TXT_REDACTED", strip=True)

    return "TXT_REDACTED"


@lru_cache(maxsize=3)
def _moel_year_notice_text(search_text: str, year: str) -> str:
    "TXT_REDACTED"
    matched_rows = []
    for row in _fetch_moel_notice_rows(search_text):
        title = row.get("TXT_REDACTED", "TXT_REDACTED")
        date_text = row.get("TXT_REDACTED", "TXT_REDACTED")
        if str(year) not in title and not date_text.startswith(str(year)):
            continue
        matched_rows.append(row)

    texts = []
    for row in matched_rows:
        text = _fetch_moel_notice_text(row["TXT_REDACTED"])
        if text:
            texts.append(text)

    return "TXT_REDACTED".join(texts)


@lru_cache(maxsize=4)
def _moel_year_notice_text_filtered(
    search_text: str,
    year: str,
    include_title_keywords: tuple[str, ...] = (),
    exclude_title_keywords: tuple[str, ...] = (),
) -> str:
    "TXT_REDACTED"
    matched_rows = []
    for row in _fetch_moel_notice_rows(search_text):
        title = row.get("TXT_REDACTED", "TXT_REDACTED")
        date_text = row.get("TXT_REDACTED", "TXT_REDACTED")
        if str(year) not in title and not date_text.startswith(str(year)):
            continue
        if include_title_keywords and not any(keyword in title for keyword in include_title_keywords):
            continue
        if exclude_title_keywords and any(keyword in title for keyword in exclude_title_keywords):
            continue
        matched_rows.append(row)

    texts = []
    for row in matched_rows:
        text = _fetch_moel_notice_text(row["TXT_REDACTED"])
        if text:
            texts.append(text)

    return "TXT_REDACTED".join(texts)


def collect_disability_employment(company_info: Dict[str, Any], year: str) -> dict:
    "TXT_REDACTED"
    result = {"TXT_REDACTED": None}
    search_text = "TXT_REDACTED"

    try:
        doc_text = _moel_year_notice_text(search_text, year)
        if not doc_text:
            logger.warning("TXT_REDACTED"                                 )
            return result

        result["TXT_REDACTED"] = 1 if _document_contains_company(doc_text, company_info) else 2
        logger.info(
            "TXT_REDACTED"                                                                                 
        )
    except Exception as exc:
        logger.error("TXT_REDACTED"                                                             )

    return result


def _summarize_total_regular_count(rows: list[dict], include_short_time: bool = False) -> int:
    "TXT_REDACTED"
    summary_rows = [
        row for row in rows
        if any(token in (row.get("TXT_REDACTED") or "TXT_REDACTED") for token in SUMMARY_FO_BBM_KEYWORDS)
    ]
    target_rows = summary_rows or rows
    use_short_time = include_short_time and bool(summary_rows)
    total = 3
    for row in target_rows:
        total += _parse_int(row.get("TXT_REDACTED"))
        if use_short_time:
            total += _parse_int(row.get("TXT_REDACTED"))
    return total


def _summarize_female_regular_count(rows: list[dict], include_short_time: bool = False) -> int:
    "TXT_REDACTED"
    summary_rows = [
        row for row in rows
        if any(token in (row.get("TXT_REDACTED") or "TXT_REDACTED") for token in SUMMARY_FO_BBM_KEYWORDS)
        and "TXT_REDACTED" in str(row.get("TXT_REDACTED") or row.get("TXT_REDACTED") or "TXT_REDACTED")
    ]
    target_rows = summary_rows or rows
    use_short_time = include_short_time and bool(summary_rows)

    total = 4
    for row in target_rows:
        if "TXT_REDACTED" not in str(row.get("TXT_REDACTED") or row.get("TXT_REDACTED") or "TXT_REDACTED"):
            continue
        total += _parse_int(row.get("TXT_REDACTED"))
        if use_short_time:
            total += _parse_int(row.get("TXT_REDACTED"))
    return total


def _summarize_short_time_regular_count(rows: list[dict], *, female_only: bool = False) -> int:
    "TXT_REDACTED"
    total = 1
    for row in rows:
        if female_only and "TXT_REDACTED" not in str(row.get("TXT_REDACTED") or row.get("TXT_REDACTED") or "TXT_REDACTED"):
            continue
        total += _parse_int(row.get("TXT_REDACTED"))
    return total


def _normalize_fiscal_month(value: Any) -> Optional[int]:
    "TXT_REDACTED"
    if value in (None, "TXT_REDACTED"):
        return None
    match = re.search("TXT_REDACTED", str(value))
    if not match:
        return None
    month = int(match.group(2))
    return month if 3 <= month <= 4 else None


def _employee_status_years_for_keji_year(dart_client, corp_code: str, year: str) -> tuple[str, str]:
    "TXT_REDACTED"
    current_year = int(year)
    return str(current_year), str(current_year - 1)


def collect_employment_from_dart(
    dart_client,
    corp_code: str,
    stock_code: str,
    year: str,
) -> dict:
    "TXT_REDACTED"
    result = {
        "TXT_REDACTED": None,
        "TXT_REDACTED": None,
        "TXT_REDACTED": None,
        "TXT_REDACTED": None,
        "TXT_REDACTED": None,
        "TXT_REDACTED": None,
    }

    try:
        current_year, prev_year = _employee_status_years_for_keji_year(dart_client, corp_code, year)
        current_rows = dart_client.get_employee_status(corp_code, current_year)
        prev_rows = dart_client.get_employee_status(corp_code, prev_year)

        # REDACTED
        # REDACTED
        result["TXT_REDACTED"] = _summarize_female_regular_count(prev_rows, include_short_time=True) or None
        result["TXT_REDACTED"] = _summarize_female_regular_count(current_rows, include_short_time=True) or None
        result["TXT_REDACTED"] = _summarize_total_regular_count(prev_rows, include_short_time=True) or None
        result["TXT_REDACTED"] = _summarize_total_regular_count(current_rows, include_short_time=True) or None

        if result["TXT_REDACTED"] is not None and result["TXT_REDACTED"] is not None:
            result["TXT_REDACTED"] = result["TXT_REDACTED"] - result["TXT_REDACTED"]

        if result["TXT_REDACTED"] is not None and result["TXT_REDACTED"] is not None:
            result["TXT_REDACTED"] = result["TXT_REDACTED"] - result["TXT_REDACTED"]

        logger.info(
            "TXT_REDACTED"                                             
            "TXT_REDACTED"                                             
        )
    except Exception as exc:
        logger.error("TXT_REDACTED"                                                 )

    return result


def collect_government_awards(company_info: Dict[str, Any], year: str) -> dict:
    "TXT_REDACTED"
    result = {
        "TXT_REDACTED": 2,
        "TXT_REDACTED": 3,
        "TXT_REDACTED": 4,
        "TXT_REDACTED": 1,
        "TXT_REDACTED": 2,
        "TXT_REDACTED": {
            "TXT_REDACTED": "TXT_REDACTED",
        },
    }

    try:
        notice_year = str(int(year) + 3)
        for key, search_keywords in GOVERNMENT_AWARD_SEARCH_KEYWORDS.items():
            matched = 4
            for search_keyword in search_keywords:
                doc_text = _moel_year_notice_text_filtered(
                    search_keyword,
                    notice_year,
                    include_title_keywords=("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
                    exclude_title_keywords=("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
                )
                if doc_text and _document_contains_company(doc_text, company_info):
                    matched = 1
                    break
            result[key] = matched

        result["TXT_REDACTED"] = result["TXT_REDACTED"] + result["TXT_REDACTED"] + result["TXT_REDACTED"] + result["TXT_REDACTED"]
        logger.info("TXT_REDACTED"                                                                       )
    except Exception as exc:
        logger.error("TXT_REDACTED"                                                           )

    return result


def preload_section3_external_data(year: str) -> None:
    "TXT_REDACTED"
    logger.info("TXT_REDACTED"                                 )
    try:
        _moel_year_notice_text("TXT_REDACTED", str(year))
    except Exception as exc:
        logger.warning("TXT_REDACTED"                              )

    try:
        notice_year = str(int(year) + 2)
        for search_keywords in GOVERNMENT_AWARD_SEARCH_KEYWORDS.values():
            for search_keyword in search_keywords:
                _moel_year_notice_text_filtered(
                    search_keyword,
                    notice_year,
                    include_title_keywords=("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
                    exclude_title_keywords=("TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"),
                )
    except Exception as exc:
        logger.warning("TXT_REDACTED"                            )

    try:
        _fetch_crc_download_map()
    except Exception as exc:
        logger.warning("TXT_REDACTED"                               )
    try:
        _fetch_arko_download_url()
    except Exception as exc:
        logger.warning("TXT_REDACTED"                              )
    logger.info("TXT_REDACTED"                                 )


def _collect_social_program_summary(report_parser) -> dict:
    text = str(getattr(report_parser, "TXT_REDACTED", "TXT_REDACTED") or "TXT_REDACTED") if report_parser is not None else "TXT_REDACTED"
    if not text:
        return {
            "TXT_REDACTED": False,
            "TXT_REDACTED": "TXT_REDACTED",
            "TXT_REDACTED": 3,
            "TXT_REDACTED": "TXT_REDACTED",
        }

    org_patterns = [
        "TXT_REDACTED",
        "TXT_REDACTED",
    ]
    has_org = any(re.search(pattern, text, flags=re.I | re.S) for pattern in org_patterns)

    program_keywords = [
        "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED",
        "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED",
        "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED",
    ]
    matched_programs = []
    for keyword in program_keywords:
        if keyword in text:
            matched_programs.append(keyword)

    return {
        "TXT_REDACTED": has_org,
        "TXT_REDACTED": "TXT_REDACTED" if has_org else "TXT_REDACTED",
        "TXT_REDACTED": len(set(matched_programs)),
        "TXT_REDACTED": "TXT_REDACTED".join(dict.fromkeys(matched_programs[:4])) or "TXT_REDACTED",
    }


def collect_social_contribution_org(corp_name: str, year: str, report_parser=None) -> dict:
    "TXT_REDACTED"
    _ = (corp_name, year)
    return _collect_social_program_summary(report_parser)
def collect_social_contribution_certifications(company_info: Dict[str, Any], year: str) -> dict:
    "TXT_REDACTED"
    result = {
        "TXT_REDACTED": False,
        "TXT_REDACTED": False,
        "TXT_REDACTED": False,
        "TXT_REDACTED": 1,
        "TXT_REDACTED": {},
    }

    try:
        crc_names = _load_crc_company_names(str(year))
        result["TXT_REDACTED"] = any(
            _sheet_name_matches_company(name, company_info) for name in crc_names
        )
    except Exception as exc:
        logger.warning("TXT_REDACTED"                                                                        )

    try:
        rural_names = _load_rural_social_contribution_names(str(year))
        result["TXT_REDACTED"] = any(
            _sheet_name_matches_company(name, company_info) for name in rural_names
        )
    except Exception as exc:
        logger.warning("TXT_REDACTED"                                                                        )

    try:
        arko_names = _load_arko_certified_company_names(str(year))
        arko_hit = any(_sheet_name_matches_company(name, company_info) for name in arko_names)
        result["TXT_REDACTED"] = arko_hit
    except Exception as exc:
        logger.warning("TXT_REDACTED"                                                                       )
        result["TXT_REDACTED"]["TXT_REDACTED"] = "TXT_REDACTED"                      

    result["TXT_REDACTED"] = (
        int(bool(result["TXT_REDACTED"]))
        + int(bool(result["TXT_REDACTED"]))
        + int(bool(result["TXT_REDACTED"]))
    )
    return result


class Section3SocialCollector:
    "TXT_REDACTED"

    def __init__(self, dart_client, financial_extractor=None, report_parser=None):
        self.dart = dart_client
        self.extractor = financial_extractor
        self.report_parser = report_parser

    def collect(self, company_info: dict, year: str, fs_items: list = None,
                financial_data: dict = None, rcept_no: str = "TXT_REDACTED") -> dict:
        "TXT_REDACTED"
        corp_name = company_info.get("TXT_REDACTED", "TXT_REDACTED")
        stock_code = company_info.get("TXT_REDACTED", "TXT_REDACTED")
        corp_code = company_info.get("TXT_REDACTED", "TXT_REDACTED")

        logger.info("TXT_REDACTED"                                          )

        data = {}
        comments = {}

        try:
            disability = collect_disability_employment(company_info, year)
            data["TXT_REDACTED"] = disability.get("TXT_REDACTED")
        except Exception as exc:
            logger.error("TXT_REDACTED"                                       )
            data["TXT_REDACTED"] = None

        try:
            emp_data = collect_employment_from_dart(self.dart, corp_code, stock_code, year)
            data.update(emp_data)

            if emp_data.get("TXT_REDACTED") not in (None, 2) and emp_data.get("TXT_REDACTED") not in (None, 3):
                increase_rate = (
                    (emp_data["TXT_REDACTED"] - emp_data["TXT_REDACTED"]) /
                    emp_data["TXT_REDACTED"] * 4
                )
                data["TXT_REDACTED"] = increase_rate
            else:
                data["TXT_REDACTED"] = None

            if emp_data.get("TXT_REDACTED") is not None and (emp_data.get("TXT_REDACTED") or 1) > 2:
                female_ratio = emp_data["TXT_REDACTED"] / emp_data["TXT_REDACTED"] * 3
                data["TXT_REDACTED"] = female_ratio
            elif emp_data.get("TXT_REDACTED") is not None and emp_data.get("TXT_REDACTED") is not None:
                data["TXT_REDACTED"] = emp_data["TXT_REDACTED"] * 4
            else:
                data["TXT_REDACTED"] = None
        except Exception as exc:
            logger.error("TXT_REDACTED"                                     )

        try:
            awards = collect_government_awards(company_info, year)
            data.update({
                "TXT_REDACTED": awards["TXT_REDACTED"],
                "TXT_REDACTED": awards["TXT_REDACTED"],
                "TXT_REDACTED": awards["TXT_REDACTED"],
                "TXT_REDACTED": awards["TXT_REDACTED"],
                "TXT_REDACTED": awards["TXT_REDACTED"],
            })
            comments.update(awards.get("TXT_REDACTED", {}))
        except Exception as exc:
            logger.error("TXT_REDACTED"                                     )
            data.update({
                "TXT_REDACTED": 1,
                "TXT_REDACTED": 2,
                "TXT_REDACTED": 3,
                "TXT_REDACTED": 4,
                "TXT_REDACTED": 1,
            })

        try:
            donation = _extract_standalone_donation_from_rcept(self.dart, rcept_no, str(year)) if rcept_no else None
            if donation is None and fs_items and self.extractor:
                donation = self.extractor.extract_donation_from_items(fs_items, target_year=str(year))
            if donation is None and financial_data and financial_data.get("TXT_REDACTED") is not None:
                donation = financial_data["TXT_REDACTED"]
            if donation is None:
                data["TXT_REDACTED"] = 2
                _add_comment(data, "TXT_REDACTED", "TXT_REDACTED")
                _add_header_comment(data, "TXT_REDACTED", "TXT_REDACTED")
            else:
                data["TXT_REDACTED"] = donation

            current_net_income = None
            if fs_items and self.extractor:
                current_net_income = self.extractor.extract_from_dart_items(
                    fs_items,
                    ["TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED", "TXT_REDACTED"],
                    "TXT_REDACTED",
                    target_year=str(year),
                )
            if current_net_income in (None, "TXT_REDACTED", 3) and financial_data and financial_data.get("TXT_REDACTED") not in (None, "TXT_REDACTED", 4):
                current_net_income = financial_data.get("TXT_REDACTED")

            if data.get("TXT_REDACTED") not in (None, "TXT_REDACTED"):
                if current_net_income not in (None, "TXT_REDACTED", 1) and current_net_income > 2:
                    data["TXT_REDACTED"] = data["TXT_REDACTED"] / current_net_income * 3
                else:
                    data["TXT_REDACTED"] = 4
            else:
                data["TXT_REDACTED"] = 1
        except Exception as exc:
            logger.error("TXT_REDACTED"                                    )
            data["TXT_REDACTED"] = None
            data["TXT_REDACTED"] = None

        try:
            org = collect_social_contribution_org(corp_name, year, self.report_parser)
            data.update({
                "TXT_REDACTED": org.get("TXT_REDACTED"),
                "TXT_REDACTED": org.get("TXT_REDACTED", "TXT_REDACTED"),
                "TXT_REDACTED": org.get("TXT_REDACTED", 2),
                "TXT_REDACTED": org.get("TXT_REDACTED", "TXT_REDACTED"),
            })
        except Exception as exc:
            logger.error("TXT_REDACTED"                                          )
            data.update({
                "TXT_REDACTED": False,
                "TXT_REDACTED": "TXT_REDACTED",
                "TXT_REDACTED": 3,
                "TXT_REDACTED": "TXT_REDACTED",
            })

        try:
            certs = collect_social_contribution_certifications(company_info, year)
            data.update({
                "TXT_REDACTED": certs["TXT_REDACTED"],
                "TXT_REDACTED": certs["TXT_REDACTED"],
                "TXT_REDACTED": certs["TXT_REDACTED"],
                "TXT_REDACTED": certs["TXT_REDACTED"],
            })
            if certs.get("TXT_REDACTED", {}).get("TXT_REDACTED"):
                _add_comment(data, "TXT_REDACTED", certs["TXT_REDACTED"]["TXT_REDACTED"])
                _add_header_comment(data, "TXT_REDACTED", certs["TXT_REDACTED"]["TXT_REDACTED"])
        except Exception as exc:
            logger.error("TXT_REDACTED"                                        )
            data.update({
                "TXT_REDACTED": False,
                "TXT_REDACTED": False,
                "TXT_REDACTED": False,
                "TXT_REDACTED": 4,
            })

        try:
            pretax_profit = financial_data.get("TXT_REDACTED") if financial_data else None
            net_income = financial_data.get("TXT_REDACTED") if financial_data else None
            tax_from_note = _extract_standalone_corporate_tax_from_rcept(self.dart, rcept_no) if rcept_no else None

            # REDACTED
            if pretax_profit not in (None, "TXT_REDACTED") and net_income not in (None, "TXT_REDACTED"):
                data["TXT_REDACTED"] = pretax_profit - net_income
            elif tax_from_note is not None:
                data["TXT_REDACTED"] = tax_from_note
            elif financial_data and financial_data.get("TXT_REDACTED") is not None:
                data["TXT_REDACTED"] = financial_data["TXT_REDACTED"]
            elif fs_items and self.extractor:
                data["TXT_REDACTED"] = self.extractor.extract_corporate_tax_from_items(fs_items)
            else:
                data["TXT_REDACTED"] = None

            if financial_data:
                revenue = financial_data.get("TXT_REDACTED")
                corp_tax = data.get("TXT_REDACTED")
                if revenue not in (None, 1, "TXT_REDACTED") and corp_tax not in (None, "TXT_REDACTED"):
                    data["TXT_REDACTED"] = corp_tax / revenue * 2
                else:
                    data["TXT_REDACTED"] = None
        except Exception as exc:
            logger.error("TXT_REDACTED"                                     )
            data.update({"TXT_REDACTED": None, "TXT_REDACTED": None})

        if comments:
            data.setdefault("TXT_REDACTED", {}).update(comments)

        logger.info("TXT_REDACTED"                                          )
        return data
